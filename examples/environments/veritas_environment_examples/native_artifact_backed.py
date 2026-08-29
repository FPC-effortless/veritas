from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook

from investigation_world.authoring import EnvironmentBuilder
from investigation_world.operational import ActionKind, WorldDomain

from ._common import execute_episode, require_perfect


def build_environment(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        sheet = workbook["Control"]
        current = str(sheet["B2"].value)
        expected = str(sheet["B3"].value)
    finally:
        workbook.close()
    return (
        EnvironmentBuilder(
            name="native-artifact-backed",
            domain=WorldDomain.FINANCIAL_SPREADSHEET,
            objective="Repair state derived from a native XLSX workbook.",
            role="spreadsheet_operator",
        )
        .system("WORKBOOK")
        .action(
            "apply_workbook_control",
            kind=ActionKind.WRITE,
            system="WORKBOOK",
            description="Apply the workbook control value.",
            parameters=("cell", "value"),
        )
        .record(
            "xlsx-rec-001",
            system="WORKBOOK",
            record_type="native_xlsx_control",
            object_id="Control!B2",
            fields={"artifact": path.name, "current": current, "expected": expected},
            searchable_text=f"workbook control {current} expected {expected}",
        )
        .initial_state(**{"Control!B2.value": current})
        .target("Control!B2", "value", expected)
        .transition(
            "apply_workbook_control",
            required_parameters={"cell": "Control!B2", "value": expected},
            set_state={"Control!B2.value": expected},
            observable_result={"accepted": True},
        )
        .require_action("apply_workbook_control")
        .require_evidence("xlsx-rec-001")
        .metadata(public={"backend": "xlsx", "native_artifact": True})
        .success("The runtime state matches the workbook control.")
        .build()
    )


def run_demo():
    with TemporaryDirectory() as directory:
        path = Path(directory) / "control.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Control"
        sheet["A2"] = "Current"
        sheet["B2"] = "draft"
        sheet["A3"] = "Expected"
        sheet["B3"] = "approved"
        workbook.save(path)
        workbook.close()
        result = execute_episode(
            build_environment(path),
            actions=(("apply_workbook_control", {"cell": "Control!B2", "value": "approved"}),),
            evidence_ids=("xlsx-rec-001",),
            claimed_state={"Control!B2.value": "approved"},
            conclusion="The native workbook control was applied.",
        )
    return require_perfect(result)
