from __future__ import annotations

import csv
import json
from datetime import date, datetime
from enum import StrEnum
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]
from pydantic import BaseModel, ConfigDict, Field, HttpUrl, model_validator

from investigation_world.foundry.models import stable_hash
from investigation_world.foundry.public_investigation_data import DatasetSplit


class StructuredInputFormat(StrEnum):
    CSV = "csv"
    JSON = "json"
    JSONL = "jsonl"
    XLSX = "xlsx"


class FieldExposure(StrEnum):
    PUBLIC = "public"
    VERIFIER = "verifier"
    IGNORE = "ignore"


class StructuredFieldRule(BaseModel):
    model_config = ConfigDict(extra="forbid")

    source_field: str
    exposure: FieldExposure
    target_field: str | None = None
    required: bool = False

    @model_validator(mode="after")
    def validate_target(self) -> StructuredFieldRule:
        if self.exposure != FieldExposure.IGNORE and self.target_field is None:
            self.target_field = self.source_field
        if self.exposure == FieldExposure.IGNORE and self.target_field is not None:
            raise ValueError("ignored fields may not define target_field")
        return self


class StructuredSourceProfile(BaseModel):
    model_config = ConfigDict(extra="forbid")

    profile_id: str
    version: str
    source_id: str
    input_format: StructuredInputFormat
    source_url: HttpUrl
    domain: str
    objective: str
    split: DatasetSplit = DatasetSplit.TRAIN_REFERENCE
    sheet_name: str | None = None
    json_records_key: str | None = None
    title_fields: list[str] = Field(min_length=1)
    location_fields: list[str] = Field(default_factory=list)
    event_date_field: str | None = None
    event_date_formats: list[str] = Field(default_factory=lambda: ["%Y-%m-%d"])
    event_year_field: str | None = None
    event_month_field: str | None = None
    field_rules: list[StructuredFieldRule] = Field(min_length=1)

    @model_validator(mode="after")
    def validate_profile(self) -> StructuredSourceProfile:
        fields = [rule.source_field for rule in self.field_rules]
        if len(fields) != len(set(fields)):
            raise ValueError("structured source field rules must be unique")
        if self.event_date_field is None and self.event_year_field is None:
            raise ValueError("event_date_field or event_year_field is required")
        classified = {rule.source_field: rule.exposure for rule in self.field_rules}
        structural = [
            *self.title_fields,
            *self.location_fields,
            self.event_date_field,
            self.event_year_field,
            self.event_month_field,
        ]
        for field_name in (item for item in structural if item is not None):
            exposure = classified.get(field_name)
            if exposure is None:
                raise ValueError(f"structural field is not classified: {field_name}")
            if exposure != FieldExposure.PUBLIC:
                raise ValueError(f"structural field must be public: {field_name}")
        if not any(rule.exposure == FieldExposure.VERIFIER for rule in self.field_rules):
            raise ValueError("structured source profile must define verifier fields")
        return self


class StructuredInvestigationCase(BaseModel):
    model_config = ConfigDict(extra="forbid")

    case_id: str
    source_id: str
    row_number: int = Field(ge=1)
    title: str
    domain: str
    event_date: date
    location: str | None = None
    split: DatasetSplit
    objective: str
    source_url: HttpUrl
    public_payload: dict[str, Any]
    verifier_payload: dict[str, Any]

    def public_projection(self) -> dict[str, Any]:
        return {
            "case_id": self.case_id,
            "source_id": self.source_id,
            "title": self.title,
            "domain": self.domain,
            "event_date": self.event_date.isoformat(),
            "location": self.location,
            "split": self.split.value,
            "objective": self.objective,
            "evidence": self.public_payload,
        }

    def verifier_projection(self) -> dict[str, Any]:
        return {
            "case_id": self.case_id,
            "source_id": self.source_id,
            "row_number": self.row_number,
            "source_url": str(self.source_url),
            "verifier": self.verifier_payload,
        }


class StructuredInvestigationCorpus(BaseModel):
    model_config = ConfigDict(extra="forbid")

    dataset_id: str
    version: str
    as_of: date
    profile_id: str
    source_id: str
    cases: list[StructuredInvestigationCase]

    def public_hash(self) -> str:
        return stable_hash([case.public_projection() for case in self.cases])

    def verifier_hash(self) -> str:
        return stable_hash([case.verifier_projection() for case in self.cases])


class StructuredCorpusWriteResult(BaseModel):
    model_config = ConfigDict(extra="forbid")

    dataset_id: str
    cases: int = Field(ge=0)
    public_output: str
    verifier_output: str
    manifest_output: str
    public_hash: str
    verifier_hash: str


def load_structured_source_profile(path: Path) -> StructuredSourceProfile:
    return StructuredSourceProfile.model_validate_json(path.read_text(encoding="utf-8"))


def _normalize_value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, dict):
        return {str(key): _normalize_value(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_normalize_value(item) for item in value]
    return str(value)


def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise ValueError("CSV input has no header row")
        if len(reader.fieldnames) != len(set(reader.fieldnames)):
            raise ValueError("CSV input contains duplicate header names")
        return [dict(row) for row in reader]


def _read_json(path: Path, records_key: str | None) -> list[dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if records_key is not None:
        if not isinstance(payload, dict):
            raise ValueError("JSON records_key requires an object root")
        payload = payload.get(records_key)
    if not isinstance(payload, list):
        raise ValueError("JSON structured input must contain a list of records")
    if not all(isinstance(row, dict) for row in payload):
        raise ValueError("JSON structured input records must be objects")
    return [dict(row) for row in payload]


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for line_number, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        if not line.strip():
            continue
        payload = json.loads(line)
        if not isinstance(payload, dict):
            raise ValueError(f"JSONL line {line_number} is not an object")
        rows.append(dict(payload))
    return rows


def _read_xlsx(path: Path, sheet_name: str | None) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name is not None else workbook.active
        if worksheet is None:
            raise ValueError("XLSX input has no active worksheet")
        iterator = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(iterator)
        except StopIteration as exc:
            raise ValueError("XLSX input has no header row") from exc
        headers = [str(value).strip() if value is not None else "" for value in header_row]
        if not headers or any(not header for header in headers):
            raise ValueError("XLSX input contains blank header cells")
        if len(headers) != len(set(headers)):
            raise ValueError("XLSX input contains duplicate header names")
        rows: list[dict[str, Any]] = []
        for values in iterator:
            rows.append(
                {
                    header: _normalize_value(value)
                    for header, value in zip(headers, values, strict=False)
                }
            )
        return rows
    finally:
        workbook.close()


def read_structured_records(
    path: Path,
    profile: StructuredSourceProfile,
) -> list[dict[str, Any]]:
    if profile.input_format == StructuredInputFormat.CSV:
        return _read_csv(path)
    if profile.input_format == StructuredInputFormat.JSON:
        return _read_json(path, profile.json_records_key)
    if profile.input_format == StructuredInputFormat.JSONL:
        return _read_jsonl(path)
    if profile.input_format == StructuredInputFormat.XLSX:
        return _read_xlsx(path, profile.sheet_name)
    raise ValueError(f"unsupported structured input format: {profile.input_format}")


def _nonempty(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def _parse_event_date(row: dict[str, Any], profile: StructuredSourceProfile) -> date:
    if profile.event_date_field is not None:
        raw = row.get(profile.event_date_field)
        if not _nonempty(raw):
            raise ValueError(f"missing event date field: {profile.event_date_field}")
        text = str(raw).strip()
        for format_string in profile.event_date_formats:
            try:
                return datetime.strptime(text, format_string).date()
            except ValueError:
                continue
        try:
            return date.fromisoformat(text[:10])
        except ValueError as exc:
            raise ValueError(f"could not parse event date {text!r}") from exc

    assert profile.event_year_field is not None
    year_raw = row.get(profile.event_year_field)
    if not _nonempty(year_raw):
        raise ValueError(f"missing event year field: {profile.event_year_field}")
    year = int(float(str(year_raw)))
    month = 1
    if profile.event_month_field is not None:
        month_raw = row.get(profile.event_month_field)
        if _nonempty(month_raw):
            month = int(float(str(month_raw)))
    if month < 1 or month > 12:
        raise ValueError(f"invalid event month: {month}")
    return date(year, month, 1)


def _joined_fields(row: dict[str, Any], fields: list[str]) -> str | None:
    values = [str(row.get(field, "")).strip() for field in fields]
    nonempty = [value for value in values if value]
    return " — ".join(nonempty) if nonempty else None


def _classify_row(
    row: dict[str, Any],
    profile: StructuredSourceProfile,
    *,
    row_number: int,
) -> StructuredInvestigationCase:
    rule_by_field = {rule.source_field: rule for rule in profile.field_rules}
    unmapped = sorted(field for field in row if field not in rule_by_field)
    if unmapped:
        raise ValueError(
            f"row {row_number} contains unclassified fields: {', '.join(unmapped)}"
        )

    public_payload: dict[str, Any] = {}
    verifier_payload: dict[str, Any] = {}
    for rule in profile.field_rules:
        value = _normalize_value(row.get(rule.source_field))
        if rule.required and not _nonempty(value):
            raise ValueError(f"row {row_number} missing required field: {rule.source_field}")
        if rule.exposure == FieldExposure.IGNORE:
            continue
        assert rule.target_field is not None
        if rule.exposure == FieldExposure.PUBLIC:
            public_payload[rule.target_field] = value
        else:
            verifier_payload[rule.target_field] = value

    title = _joined_fields(row, profile.title_fields)
    if title is None:
        raise ValueError(f"row {row_number} has no usable title fields")
    location = _joined_fields(row, profile.location_fields)
    case_hash = stable_hash(
        {"source_id": profile.source_id, "row_number": row_number, "public": public_payload}
    )[:16]
    return StructuredInvestigationCase(
        case_id=f"{profile.source_id}-{case_hash}",
        source_id=profile.source_id,
        row_number=row_number,
        title=title,
        domain=profile.domain,
        event_date=_parse_event_date(row, profile),
        location=location,
        split=profile.split,
        objective=profile.objective,
        source_url=profile.source_url,
        public_payload=public_payload,
        verifier_payload=verifier_payload,
    )


def compile_structured_investigation_corpus(
    profile: StructuredSourceProfile,
    input_path: Path,
    *,
    dataset_id: str,
    version: str,
    as_of: date,
) -> StructuredInvestigationCorpus:
    rows = read_structured_records(input_path, profile)
    cases = [
        _classify_row(row, profile, row_number=index)
        for index, row in enumerate(rows, start=1)
    ]
    return StructuredInvestigationCorpus(
        dataset_id=dataset_id,
        version=version,
        as_of=as_of,
        profile_id=profile.profile_id,
        source_id=profile.source_id,
        cases=cases,
    )


def audit_structured_investigation_corpus(
    corpus: StructuredInvestigationCorpus,
    profile: StructuredSourceProfile,
) -> dict[str, Any]:
    verifier_targets = {
        rule.target_field
        for rule in profile.field_rules
        if rule.exposure == FieldExposure.VERIFIER and rule.target_field is not None
    }
    public_targets = {
        rule.target_field
        for rule in profile.field_rules
        if rule.exposure == FieldExposure.PUBLIC and rule.target_field is not None
    }
    overlap = sorted(public_targets & verifier_targets)
    duplicate_ids = len(corpus.cases) - len({case.case_id for case in corpus.cases})
    leaked_cases = [
        case.case_id
        for case in corpus.cases
        if verifier_targets & set(case.public_payload)
    ]
    public_serialized = json.dumps(
        [case.public_projection() for case in corpus.cases],
        sort_keys=True,
    )
    raw_source_exposed = str(profile.source_url) in public_serialized
    row_number_exposed = any("row_number" in case.public_projection() for case in corpus.cases)
    passed = (
        not overlap
        and duplicate_ids == 0
        and not leaked_cases
        and not raw_source_exposed
        and not row_number_exposed
    )
    return {
        "passed": passed,
        "cases": len(corpus.cases),
        "public_fields": sorted(public_targets),
        "verifier_fields": sorted(verifier_targets),
        "field_overlap": overlap,
        "duplicate_case_ids": duplicate_ids,
        "leaked_case_ids": leaked_cases,
        "raw_source_url_exposed": raw_source_exposed,
        "row_number_exposed": row_number_exposed,
        "public_hash": corpus.public_hash(),
        "verifier_hash": corpus.verifier_hash(),
    }


def write_structured_investigation_corpus(
    corpus: StructuredInvestigationCorpus,
    profile: StructuredSourceProfile,
    *,
    public_output: Path,
    verifier_output: Path,
    manifest_output: Path,
) -> StructuredCorpusWriteResult:
    audit = audit_structured_investigation_corpus(corpus, profile)
    if not audit["passed"]:
        raise ValueError(f"structured corpus boundary audit failed: {audit}")

    public_output.parent.mkdir(parents=True, exist_ok=True)
    verifier_output.parent.mkdir(parents=True, exist_ok=True)
    manifest_output.parent.mkdir(parents=True, exist_ok=True)
    public_output.write_text(
        "".join(
            json.dumps(case.public_projection(), sort_keys=True) + "\n"
            for case in corpus.cases
        ),
        encoding="utf-8",
    )
    verifier_output.write_text(
        "".join(
            json.dumps(case.verifier_projection(), sort_keys=True) + "\n"
            for case in corpus.cases
        ),
        encoding="utf-8",
    )
    public_audit = {
        "passed": audit["passed"],
        "cases": audit["cases"],
        "field_overlap": audit["field_overlap"],
        "duplicate_case_ids": audit["duplicate_case_ids"],
        "leaked_case_count": len(audit["leaked_case_ids"]),
        "raw_source_url_exposed": audit["raw_source_url_exposed"],
        "row_number_exposed": audit["row_number_exposed"],
        "public_hash": audit["public_hash"],
    }
    manifest = {
        "dataset_id": corpus.dataset_id,
        "version": corpus.version,
        "as_of": corpus.as_of.isoformat(),
        "profile_id": corpus.profile_id,
        "source_id": corpus.source_id,
        "cases": len(corpus.cases),
        "split": profile.split.value,
        "audit": public_audit,
        "public_output": str(public_output),
        "verifier_materialized": True,
    }
    manifest_output.write_text(
        json.dumps(manifest, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    return StructuredCorpusWriteResult(
        dataset_id=corpus.dataset_id,
        cases=len(corpus.cases),
        public_output=str(public_output),
        verifier_output=str(verifier_output),
        manifest_output=str(manifest_output),
        public_hash=corpus.public_hash(),
        verifier_hash=corpus.verifier_hash(),
    )
