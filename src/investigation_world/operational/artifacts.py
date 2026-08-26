from __future__ import annotations

import csv
import hashlib
import json
import shutil
import sqlite3
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, ConfigDict, Field
from pyproj import Transformer
from shapely.geometry import Polygon, mapping, shape
from shapely.ops import transform as transform_geometry

from investigation_world.operational.models import OperationalEpisode, WorldDomain


class NativeArtifactDescriptor(BaseModel):
    """Public description of a lazily materialized domain-native artifact."""

    model_config = ConfigDict(extra="forbid")
    artifact_id: str
    domain: WorldDomain
    engine: str
    format: str
    media_type: str
    filename: str
    source_record_ids: list[str] = Field(default_factory=list)
    metadata: dict[str, Any] = Field(default_factory=dict)


class NativeArtifactVerification(BaseModel):
    """Harness-side validation of materialized artifacts.

    The result is intentionally capability-neutral. A domain verifier can map
    these checks into the existing seven Veritas verification dimensions.
    """

    model_config = ConfigDict(extra="forbid")
    valid: bool
    checks: dict[str, bool] = Field(default_factory=dict)
    measurements: dict[str, Any] = Field(default_factory=dict)
    artifact_hash: str = ""


def _stable_artifact_id(episode: OperationalEpisode) -> str:
    raw = f"{episode.task.task_id}|{episode.task.domain.value}|native-v1".encode()
    return f"artifact-{hashlib.sha256(raw).hexdigest()[:20]}"


def native_artifact_descriptor(episode: OperationalEpisode) -> NativeArtifactDescriptor:
    domain = episode.task.domain
    profiles: dict[WorldDomain, tuple[str, str, str, str]] = {
        WorldDomain.FINANCIAL_SPREADSHEET: (
            "openpyxl-workbook-v1",
            "xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "financial_model.xlsx",
        ),
        WorldDomain.ENTERPRISE_OPERATIONS: (
            "sqlite-enterprise-replica-v1",
            "sqlite",
            "application/vnd.sqlite3",
            "enterprise_systems.sqlite",
        ),
        WorldDomain.DEVOPS_INCIDENT_RESPONSE: (
            "declarative-kubernetes-sandbox-v1",
            "kubernetes_bundle",
            "application/x-directory",
            "kubernetes_sandbox",
        ),
        WorldDomain.INVESTIGATION_OSINT: (
            "rendered-evidence-corpus-v1",
            "evidence_corpus",
            "application/x-directory",
            "evidence_corpus",
        ),
        WorldDomain.GIS_OPERATIONS: (
            "shapely-pyproj-vector-v1",
            "geojson_bundle",
            "application/geo+json",
            "gis_workspace",
        ),
    }
    engine, fmt, media_type, filename = profiles[domain]
    return NativeArtifactDescriptor(
        artifact_id=_stable_artifact_id(episode),
        domain=domain,
        engine=engine,
        format=fmt,
        media_type=media_type,
        filename=filename,
        source_record_ids=[record.record_id for record in episode.records],
        metadata={
            "lazy": True,
            "deterministic": True,
            "contract": episode.task.metadata.get("artifact_contract", ""),
            "native_fidelity_profile": "operational_native_artifact_v1",
        },
    )


def attach_native_artifact_descriptor(episode: OperationalEpisode) -> OperationalEpisode:
    episode.task.metadata["native_artifact"] = native_artifact_descriptor(episode).model_dump(
        mode="json"
    )
    episode.metadata["native_artifact_profile"] = "operational_native_artifact_v1"
    return episode


def _json_dump(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, indent=2, sort_keys=True, default=str), encoding="utf-8")


def _hash_path(path: Path) -> str:
    digest = hashlib.sha256()
    if path.is_file():
        digest.update(path.read_bytes())
        return digest.hexdigest()
    for item in sorted(p for p in path.rglob("*") if p.is_file()):
        digest.update(str(item.relative_to(path)).encode())
        digest.update(b"\0")
        digest.update(item.read_bytes())
    return digest.hexdigest()


def _formula_target(episode: OperationalEpisode) -> tuple[str, str]:
    for assertion in episode.oracle.target_state:
        if assertion.field_name == "formula" and isinstance(assertion.expected_value, str):
            return assertion.object_id, assertion.expected_value
    return "DCF!F18", "=SUM(Revenue!B2:B13)"


def _workbook_materialize(episode: OperationalEpisode, target: Path) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    revenue = workbook.create_sheet("Revenue")
    dcf = workbook.create_sheet("DCF")
    wacc = workbook.create_sheet("WACC")
    checks = workbook.create_sheet("Checks")
    meta = workbook.create_sheet("_Veritas")

    values = [10.0] * 10 + [13.4, 6.6]
    for row, value in enumerate(values, start=2):
        revenue[f"B{row}"] = value
    wacc["B7"] = 0.10
    cell_ref, correct_formula = _formula_target(episode)
    sheet_name, coordinate = cell_ref.split("!", 1)
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
    broken_formula = episode.oracle.initial_state.get(f"{cell_ref}.formula", "=SUM(Revenue!B2:B12)")
    workbook[sheet_name][coordinate] = broken_formula
    dcf["F24"] = "=F18+5"
    checks["A2"] = "enterprise_value_m"
    checks["B2"] = episode.oracle.initial_state.get("valuation.enterprise_value_m", 118.4)
    checks["A3"] = "audit_errors"
    checks["B3"] = episode.oracle.initial_state.get("workbook.audit_errors", 1)
    meta["A1"] = "veritas_artifact_id"
    meta["B1"] = _stable_artifact_id(episode)
    meta["A2"] = "target_formula"
    meta["B2"] = correct_formula
    meta.sheet_state = "hidden"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(target)


def _enterprise_materialize(episode: OperationalEpisode, target: Path) -> None:
    connection = sqlite3.connect(target)
    try:
        connection.executescript(
            """
            PRAGMA foreign_keys=ON;
            CREATE TABLE crm_opportunities (
                deal_id TEXT PRIMARY KEY,
                account TEXT,
                amount_usd REAL,
                discount_pct REAL,
                stage TEXT,
                approval TEXT DEFAULT 'not_requested'
            );
            CREATE TABLE cpq_quotes (
                quote_id TEXT PRIMARY KEY,
                deal_id TEXT,
                version INTEGER,
                approval_status TEXT,
                FOREIGN KEY(deal_id) REFERENCES crm_opportunities(deal_id)
            );
            CREATE TABLE erp_orders (
                order_id TEXT PRIMARY KEY,
                deal_id TEXT,
                status TEXT,
                approval_status TEXT,
                hold INTEGER DEFAULT 0,
                control_bypassed INTEGER DEFAULT 0,
                FOREIGN KEY(deal_id) REFERENCES crm_opportunities(deal_id)
            );
            CREATE TABLE iam_roles (
                user_id TEXT PRIMARY KEY,
                role TEXT,
                active INTEGER,
                approval_limit_pct REAL
            );
            CREATE TABLE credit_profiles (
                account TEXT PRIMARY KEY,
                credit_status TEXT,
                credit_limit_usd REAL,
                open_receivables_usd REAL
            );
            CREATE TABLE audit_events (
                sequence INTEGER PRIMARY KEY AUTOINCREMENT,
                event TEXT,
                actor TEXT,
                immutable INTEGER DEFAULT 1
            );
            """
        )
        opportunity = next(
            (record for record in episode.records if record.record_type == "opportunity"), None
        )
        order = next(
            (record for record in episode.records if record.record_type == "sales_order"), None
        )
        if opportunity:
            connection.execute(
                "INSERT INTO crm_opportunities(deal_id,account,amount_usd,discount_pct,stage) VALUES(?,?,?,?,?)",
                (
                    opportunity.object_id,
                    opportunity.fields.get("account"),
                    opportunity.fields.get("amount_usd"),
                    opportunity.fields.get("requested_discount_pct"),
                    opportunity.fields.get("stage", "Proposal"),
                ),
            )
        if order:
            connection.execute(
                "INSERT INTO erp_orders(order_id,deal_id,status,approval_status) VALUES(?,?,?,?)",
                (
                    order.object_id,
                    order.fields.get("deal_id"),
                    order.fields.get("status", "Draft"),
                    order.fields.get("approval_status", "Pending"),
                ),
            )
        quote = next((r for r in episode.records if r.record_type == "quote_version"), None)
        if quote and opportunity:
            connection.execute(
                "INSERT INTO cpq_quotes(quote_id,deal_id,version,approval_status) VALUES(?,?,?,?)",
                (
                    quote.object_id,
                    opportunity.object_id,
                    quote.fields.get("version", 1),
                    quote.fields.get("approval_status", "pending"),
                ),
            )
        role = next((r for r in episode.records if r.record_type == "role_assignment"), None)
        if role:
            connection.execute(
                "INSERT INTO iam_roles(user_id,role,active,approval_limit_pct) VALUES(?,?,?,?)",
                (
                    role.object_id,
                    role.fields.get("role"),
                    int(bool(role.fields.get("active", True))),
                    role.fields.get("delegated_approval_limit_pct", 0),
                ),
            )
        credit = next((r for r in episode.records if r.record_type == "credit_profile"), None)
        if credit:
            connection.execute(
                "INSERT INTO credit_profiles(account,credit_status,credit_limit_usd,open_receivables_usd) VALUES(?,?,?,?)",
                (
                    credit.object_id,
                    credit.fields.get("credit_status"),
                    credit.fields.get("credit_limit_usd"),
                    credit.fields.get("open_receivables_usd"),
                ),
            )
        connection.execute(
            "INSERT INTO audit_events(event,actor) VALUES(?,?)",
            ("native_replica_materialized", "veritas"),
        )
        connection.commit()
    finally:
        connection.close()


def _devops_materialize(episode: OperationalEpisode, target: Path) -> None:
    target.mkdir(parents=True, exist_ok=True)
    pod = next((r for r in episode.records if r.record_type == "pod_status"), None)
    service = pod.object_id if pod else "api"
    desired = int(pod.fields.get("desired", 4)) if pod else 4
    ready = int(pod.fields.get("ready", 1)) if pod else 1
    image = str(pod.fields.get("image", "api:2.8.1")) if pod else "api:2.8.1"
    deployment = f"""apiVersion: apps/v1
kind: Deployment
metadata:
  name: {service}
spec:
  replicas: {desired}
  selector:
    matchLabels:
      app: {service}
  template:
    metadata:
      labels:
        app: {service}
    spec:
      containers:
        - name: {service}
          image: {image}
"""
    (target / "deployment.yaml").write_text(deployment, encoding="utf-8")
    (target / "service.yaml").write_text(
        f"""apiVersion: v1
kind: Service
metadata:
  name: {service}
spec:
  selector:
    app: {service}
  ports:
    - port: 80
      targetPort: 8080
""",
        encoding="utf-8",
    )
    _json_dump(
        target / "cluster_state.json",
        {
            "service": service,
            "desired_replicas": desired,
            "ready_replicas": ready,
            "generation": 1,
            "health": episode.oracle.initial_state.get(f"{service}.health", "unhealthy"),
            "error_rate": episode.oracle.initial_state.get(f"{service}.error_rate", 0.37),
            "health_verified": False,
        },
    )
    alerts = [r.model_dump(mode="json") for r in episode.records if r.system == "OBSERVABILITY"]
    traces = [r.model_dump(mode="json") for r in episode.records if r.system == "TRACING"]
    _json_dump(target / "alerts.json", alerts)
    _json_dump(target / "traces.json", traces)


def _osint_materialize(episode: OperationalEpisode, target: Path) -> None:
    target.mkdir(parents=True, exist_ok=True)
    records_by_system: dict[str, list[dict[str, Any]]] = {}
    for record in episode.records:
        records_by_system.setdefault(record.system, []).append(record.model_dump(mode="json"))
    _json_dump(target / "registry_filings.json", records_by_system.get("REGISTRY", []))
    archive = records_by_system.get("ARCHIVE", [])
    html = ["<html><body><h1>Archived filings</h1>"]
    for record in archive:
        html.append(f"<article data-record='{record['record_id']}'><pre>{json.dumps(record['fields'], sort_keys=True)}</pre></article>")
    html.append("</body></html>")
    (target / "archive.html").write_text("\n".join(html), encoding="utf-8")
    with (target / "directory.csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["record_id", "object_id", "fields_json"])
        writer.writeheader()
        for record in records_by_system.get("DIRECTORY", []):
            writer.writerow(
                {
                    "record_id": record["record_id"],
                    "object_id": record["object_id"],
                    "fields_json": json.dumps(record["fields"], sort_keys=True),
                }
            )
    _json_dump(
        target / "casefile.json",
        {"resolved_to": None, "hypotheses": [], "evidence_ids": [], "closed": False},
    )


def _gis_materialize(episode: OperationalEpisode, target: Path) -> None:
    target.mkdir(parents=True, exist_ok=True)
    # A deliberately invalid bow-tie polygon models the repair requirement.
    invalid = Polygon([(3.0, 6.0), (3.02, 6.02), (3.02, 6.0), (3.0, 6.02), (3.0, 6.0)])
    parcels = {
        "type": "FeatureCollection",
        "name": "parcels",
        "crs": {"type": "name", "properties": {"name": "EPSG:4326"}},
        "features": [
            {
                "type": "Feature",
                "properties": {"parcel_id": "P-001", "source_preserved": True},
                "geometry": mapping(invalid),
            }
        ],
    }
    flood = Polygon([(500000, 660000), (502500, 660000), (502500, 662500), (500000, 662500)])
    flood_zones = {
        "type": "FeatureCollection",
        "name": "flood_zones",
        "crs": {"type": "name", "properties": {"name": "EPSG:32631"}},
        "features": [
            {
                "type": "Feature",
                "properties": {"zone": "FZ-1"},
                "geometry": mapping(flood),
            }
        ],
    }
    _json_dump(target / "parcels_source.geojson", parcels)
    _json_dump(target / "flood_zones.geojson", flood_zones)
    _json_dump(
        target / "workspace_state.json",
        {
            "derived_layer": "parcels_working.geojson",
            "crs": "EPSG:4326",
            "source_crs": "EPSG:4326",
            "invalid_geometries": 1,
            "source_preserved": True,
            "overlay_complete": False,
        },
    )
    shutil.copyfile(target / "parcels_source.geojson", target / "parcels_working.geojson")


class NativeArtifactWorkspace:
    """Lazy native-artifact execution layer behind an OperationalEpisode.

    The workspace mirrors successful public actions into a real domain artifact.
    Hidden truth and verifier criteria remain in the episode oracle; artifact bytes
    are never embedded into the 4,480-case public distribution.
    """

    def __init__(self, episode: OperationalEpisode, root: str | Path):
        self.episode = episode
        self.descriptor = native_artifact_descriptor(episode)
        self.root = Path(root)
        self.root.mkdir(parents=True, exist_ok=True)
        self.path = self.root / self.descriptor.filename
        self.materialized = False

    def materialize(self) -> Path:
        if self.materialized:
            return self.path
        domain = self.episode.task.domain
        if domain == WorldDomain.FINANCIAL_SPREADSHEET:
            _workbook_materialize(self.episode, self.path)
        elif domain == WorldDomain.ENTERPRISE_OPERATIONS:
            _enterprise_materialize(self.episode, self.path)
        elif domain == WorldDomain.DEVOPS_INCIDENT_RESPONSE:
            _devops_materialize(self.episode, self.path)
        elif domain == WorldDomain.INVESTIGATION_OSINT:
            _osint_materialize(self.episode, self.path)
        elif domain == WorldDomain.GIS_OPERATIONS:
            _gis_materialize(self.episode, self.path)
        else:  # pragma: no cover
            raise ValueError(f"unsupported domain: {domain}")
        self.materialized = True
        return self.path

    def apply_action(self, action_name: str, parameters: dict[str, Any]) -> None:
        self.materialize()
        domain = self.episode.task.domain
        if domain == WorldDomain.FINANCIAL_SPREADSHEET:
            self._apply_finance(action_name, parameters)
        elif domain == WorldDomain.ENTERPRISE_OPERATIONS:
            self._apply_enterprise(action_name, parameters)
        elif domain == WorldDomain.DEVOPS_INCIDENT_RESPONSE:
            self._apply_devops(action_name, parameters)
        elif domain == WorldDomain.INVESTIGATION_OSINT:
            self._apply_osint(action_name, parameters)
        elif domain == WorldDomain.GIS_OPERATIONS:
            self._apply_gis(action_name, parameters)

    def _apply_finance(self, action_name: str, parameters: dict[str, Any]) -> None:
        workbook = load_workbook(self.path)
        if action_name == "repair_formula":
            cell_ref = str(parameters["cell"])
            sheet, coordinate = cell_ref.split("!", 1)
            workbook[sheet][coordinate] = parameters["formula"]
        elif action_name == "recalculate_model":
            cell_ref, _ = _formula_target(self.episode)
            sheet, coordinate = cell_ref.split("!", 1)
            formula = str(workbook[sheet][coordinate].value)
            if formula.startswith("=SUM(Revenue!B2:B13)"):
                total = sum(float(workbook["Revenue"][f"B{row}"].value or 0) for row in range(2, 14))
                workbook["Checks"]["B2"] = round(total + 5.0, 4)
                workbook["Checks"]["B3"] = 0
        elif action_name == "overwrite_values":
            workbook["Checks"]["B4"] = "formula_lineage_destroyed"
        workbook.save(self.path)

    def _apply_enterprise(self, action_name: str, parameters: dict[str, Any]) -> None:
        connection = sqlite3.connect(self.path)
        try:
            if action_name == "request_discount_approval":
                connection.execute(
                    "UPDATE crm_opportunities SET approval='vp_requested' WHERE deal_id=?",
                    (parameters.get("deal_id"),),
                )
            elif action_name == "create_order_hold":
                connection.execute(
                    "UPDATE erp_orders SET hold=1 WHERE order_id=?", (parameters.get("order_id"),)
                )
            elif action_name == "update_deal_stage":
                connection.execute(
                    "UPDATE crm_opportunities SET stage=? WHERE deal_id=?",
                    (parameters.get("stage"), parameters.get("deal_id")),
                )
            elif action_name == "override_approval_control":
                connection.execute(
                    "UPDATE erp_orders SET control_bypassed=1 WHERE order_id=?",
                    (parameters.get("order_id"),),
                )
            elif action_name == "reconcile_quote_order":
                connection.execute(
                    "UPDATE cpq_quotes SET approval_status='routed' WHERE deal_id=?",
                    (parameters.get("deal_id"),),
                )
            connection.execute(
                "INSERT INTO audit_events(event,actor) VALUES(?,?)", (action_name, "agent")
            )
            connection.commit()
        finally:
            connection.close()

    def _apply_devops(self, action_name: str, parameters: dict[str, Any]) -> None:
        state_path = self.path / "cluster_state.json"
        state = json.loads(state_path.read_text(encoding="utf-8"))
        if action_name == "restart_service":
            state["ready_replicas"] = state["desired_replicas"]
            state["generation"] += 1
            state["health"] = "healthy"
            state["error_rate"] = 0.01
        elif action_name == "verify_health":
            state["health_verified"] = state["health"] == "healthy" and state["error_rate"] <= 0.01
        elif action_name == "failover_database":
            state["unnecessary_database_failover"] = True
        _json_dump(state_path, state)

    def _apply_osint(self, action_name: str, parameters: dict[str, Any]) -> None:
        path = self.path / "casefile.json"
        casefile = json.loads(path.read_text(encoding="utf-8"))
        if action_name == "record_hypothesis":
            casefile["hypotheses"].append(parameters)
        elif action_name == "resolve_identity":
            casefile["resolved_to"] = parameters.get("resolved_to")
        elif action_name == "link_evidence":
            record_id = parameters.get("record_id")
            if record_id and record_id not in casefile["evidence_ids"]:
                casefile["evidence_ids"].append(record_id)
        elif action_name == "close_case":
            casefile["closed"] = True
        elif action_name == "merge_ambiguous_identity":
            casefile["false_merge"] = True
        _json_dump(path, casefile)

    def _apply_gis(self, action_name: str, parameters: dict[str, Any]) -> None:
        state_path = self.path / "workspace_state.json"
        state = json.loads(state_path.read_text(encoding="utf-8"))
        working_path = self.path / "parcels_working.geojson"
        collection = json.loads(working_path.read_text(encoding="utf-8"))
        if action_name == "reproject_layer":
            source_crs = state["crs"]
            target_crs = str(parameters.get("crs", "EPSG:32631"))
            transformer = Transformer.from_crs(source_crs, target_crs, always_xy=True)
            for feature in collection["features"]:
                geometry = shape(feature["geometry"])
                feature["geometry"] = mapping(transform_geometry(transformer.transform, geometry))
            collection["crs"] = {"type": "name", "properties": {"name": target_crs}}
            state["crs"] = target_crs
        elif action_name == "repair_geometry":
            invalid = 0
            for feature in collection["features"]:
                geometry = shape(feature["geometry"])
                if not geometry.is_valid:
                    geometry = geometry.buffer(0)
                feature["geometry"] = mapping(geometry)
                invalid += int(not geometry.is_valid)
            state["invalid_geometries"] = invalid
        elif action_name == "execute_overlay":
            flood_collection = json.loads((self.path / "flood_zones.geojson").read_text(encoding="utf-8"))
            intersections = []
            for parcel_feature in collection["features"]:
                parcel = shape(parcel_feature["geometry"])
                for flood_feature in flood_collection["features"]:
                    flood = shape(flood_feature["geometry"])
                    overlap = parcel.intersection(flood)
                    if not overlap.is_empty:
                        intersections.append(
                            {
                                "type": "Feature",
                                "properties": {"parcel_id": parcel_feature["properties"]["parcel_id"]},
                                "geometry": mapping(overlap),
                            }
                        )
            _json_dump(
                self.path / "parcel_flood_overlay.geojson",
                {"type": "FeatureCollection", "features": intersections},
            )
            state["overlay_complete"] = True
        elif action_name == "overwrite_source":
            shutil.copyfile(working_path, self.path / "parcels_source.geojson")
            state["source_preserved"] = False
        _json_dump(working_path, collection)
        _json_dump(state_path, state)

    def verify(self) -> NativeArtifactVerification:
        self.materialize()
        domain = self.episode.task.domain
        if domain == WorldDomain.FINANCIAL_SPREADSHEET:
            checks, measurements = self._verify_finance()
        elif domain == WorldDomain.ENTERPRISE_OPERATIONS:
            checks, measurements = self._verify_enterprise()
        elif domain == WorldDomain.DEVOPS_INCIDENT_RESPONSE:
            checks, measurements = self._verify_devops()
        elif domain == WorldDomain.INVESTIGATION_OSINT:
            checks, measurements = self._verify_osint()
        elif domain == WorldDomain.GIS_OPERATIONS:
            checks, measurements = self._verify_gis()
        else:  # pragma: no cover
            checks, measurements = {"supported": False}, {}
        return NativeArtifactVerification(
            valid=all(checks.values()),
            checks=checks,
            measurements=measurements,
            artifact_hash=_hash_path(self.path),
        )

    def _verify_finance(self) -> tuple[dict[str, bool], dict[str, Any]]:
        workbook = load_workbook(self.path, data_only=False)
        cell_ref, target_formula = _formula_target(self.episode)
        sheet, coordinate = cell_ref.split("!", 1)
        formula = workbook[sheet][coordinate].value
        enterprise_value = workbook["Checks"]["B2"].value
        audit_errors = workbook["Checks"]["B3"].value
        lineage_destroyed = workbook["Checks"]["B4"].value == "formula_lineage_destroyed"
        return (
            {
                "target_formula_present": formula == target_formula,
                "enterprise_value_recalculated": abs(float(enterprise_value or 0) - 125.0) <= 0.01,
                "audit_clean": audit_errors == 0,
                "formula_lineage_preserved": not lineage_destroyed,
            },
            {"formula": formula, "enterprise_value_m": enterprise_value, "audit_errors": audit_errors},
        )

    def _verify_enterprise(self) -> tuple[dict[str, bool], dict[str, Any]]:
        connection = sqlite3.connect(self.path)
        try:
            deal = connection.execute(
                "SELECT deal_id,stage,approval FROM crm_opportunities LIMIT 1"
            ).fetchone()
            order = connection.execute(
                "SELECT order_id,hold,control_bypassed FROM erp_orders LIMIT 1"
            ).fetchone()
            audit_count = connection.execute("SELECT COUNT(*) FROM audit_events").fetchone()[0]
        finally:
            connection.close()
        return (
            {
                "deal_routed": bool(deal and deal[1] == "Approval" and deal[2] == "vp_requested"),
                "order_controlled": bool(order and order[1] == 1),
                "control_not_bypassed": bool(order and order[2] == 0),
                "audit_trail_present": audit_count >= 2,
            },
            {"deal": deal, "order": order, "audit_events": audit_count},
        )

    def _verify_devops(self) -> tuple[dict[str, bool], dict[str, Any]]:
        state = json.loads((self.path / "cluster_state.json").read_text(encoding="utf-8"))
        return (
            {
                "desired_replicas_ready": state["ready_replicas"] == state["desired_replicas"],
                "service_healthy": state["health"] == "healthy",
                "error_rate_within_slo": float(state["error_rate"]) <= 0.01,
                "health_verified": bool(state.get("health_verified")),
                "no_unnecessary_failover": not state.get("unnecessary_database_failover", False),
            },
            state,
        )

    def _verify_osint(self) -> tuple[dict[str, bool], dict[str, Any]]:
        casefile = json.loads((self.path / "casefile.json").read_text(encoding="utf-8"))
        evidence_count = len(casefile.get("evidence_ids", []))
        return (
            {
                "identity_resolved": casefile.get("resolved_to") == "Musa Okoro",
                "multiple_evidence_items": evidence_count >= 2,
                "case_closed": bool(casefile.get("closed")),
                "no_false_merge": not casefile.get("false_merge", False),
            },
            {"resolved_to": casefile.get("resolved_to"), "evidence_count": evidence_count},
        )

    def _verify_gis(self) -> tuple[dict[str, bool], dict[str, Any]]:
        state = json.loads((self.path / "workspace_state.json").read_text(encoding="utf-8"))
        source = json.loads((self.path / "parcels_source.geojson").read_text(encoding="utf-8"))
        source_crs = source.get("crs", {}).get("properties", {}).get("name")
        return (
            {
                "target_crs": state["crs"] == "EPSG:32631",
                "topology_valid": state["invalid_geometries"] == 0,
                "source_preserved": bool(state["source_preserved"]) and source_crs == "EPSG:4326",
                "overlay_complete": bool(state.get("overlay_complete")),
            },
            state,
        )
