from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from investigation_world.operational.artifacts import NativeArtifactWorkspace, _json_dump
from investigation_world.operational.models import OperationalEpisode, StateAssertion, WorldDomain

_SUM_RANGE = re.compile(r"^=SUM\(([^!]+)!([A-Z]+)(\d+):([A-Z]+)(\d+)\)$", re.IGNORECASE)


def _target_assertion(
    episode: OperationalEpisode,
    *,
    object_id: str | None = None,
    field_name: str,
) -> StateAssertion | None:
    for assertion in episode.oracle.target_state:
        if assertion.field_name != field_name:
            continue
        if object_id is None or assertion.object_id == object_id:
            return assertion
    return None


def _finance_targets(episode: OperationalEpisode) -> tuple[str, str, float, float]:
    formula = next(
        (
            assertion
            for assertion in episode.oracle.target_state
            if assertion.field_name == "formula" and isinstance(assertion.expected_value, str)
        ),
        None,
    )
    valuation = _target_assertion(
        episode,
        object_id="valuation",
        field_name="enterprise_value_m",
    )
    if formula is None or valuation is None:
        raise ValueError("financial native artifact requires formula and valuation targets")
    initial_ev = float(
        episode.oracle.initial_state.get("valuation.enterprise_value_m", valuation.expected_value)
    )
    return (
        formula.object_id,
        str(formula.expected_value),
        float(valuation.expected_value),
        initial_ev,
    )


def _sum_formula(workbook, formula: str) -> float:
    match = _SUM_RANGE.match(formula)
    if match is None:
        raise ValueError(f"native workbook engine does not support formula: {formula}")
    sheet, start_col, start_row, end_col, end_row = match.groups()
    if start_col.upper() != end_col.upper():
        raise ValueError(f"native workbook engine only supports single-column SUM ranges: {formula}")
    return sum(
        float(workbook[sheet][f"{start_col.upper()}{row}"].value or 0.0)
        for row in range(int(start_row), int(end_row) + 1)
    )


def _devops_target(episode: OperationalEpisode) -> tuple[str, float, float]:
    assertion = next(
        (
            item
            for item in episode.oracle.target_state
            if item.field_name == "error_rate" and isinstance(item.expected_value, (int, float))
        ),
        None,
    )
    if assertion is None:
        raise ValueError("DevOps native artifact requires an error-rate target")
    return assertion.object_id, float(assertion.expected_value), float(assertion.tolerance or 0.0)


def _osint_target(episode: OperationalEpisode) -> str:
    assertion = _target_assertion(
        episode,
        object_id="investigation",
        field_name="resolved_to",
    )
    if assertion is None:
        raise ValueError("OSINT native artifact requires a resolved identity target")
    return str(assertion.expected_value)


def _gis_targets(episode: OperationalEpisode) -> tuple[str, str, str]:
    assertion = next(
        (
            item
            for item in episode.oracle.target_state
            if item.field_name == "crs" and isinstance(item.expected_value, str)
        ),
        None,
    )
    if assertion is None:
        raise ValueError("GIS native artifact requires a CRS target")
    layer = assertion.object_id
    source = str(episode.oracle.initial_state.get(f"{layer}.crs", "EPSG:4326"))
    return layer, source, str(assertion.expected_value)


class ParameterizedNativeArtifactWorkspace(NativeArtifactWorkspace):
    """Native artifact workspace whose bytes follow the case-specific oracle targets.

    The base workspace provides actual file/database/geospatial artifacts. This
    adapter removes remaining reference-case constants so any parameterized
    operational episode can be lazily materialized and verified with the same
    target values that drive its hidden state verifier.
    """

    def __init__(self, episode: OperationalEpisode, root: str | Path):
        super().__init__(episode, root)
        self._parameterized = False

    def materialize(self) -> Path:
        path = super().materialize()
        if self._parameterized:
            return path
        if self.episode.task.domain == WorldDomain.FINANCIAL_SPREADSHEET:
            self._parameterize_finance()
        elif self.episode.task.domain == WorldDomain.GIS_OPERATIONS:
            self._parameterize_gis()
        self._parameterized = True
        return path

    def _parameterize_finance(self) -> None:
        cell_ref, target_formula, target_ev, initial_ev = _finance_targets(self.episode)
        workbook = load_workbook(self.path)
        match = _SUM_RANGE.match(target_formula)
        if match is None:
            raise ValueError(f"unsupported generated finance formula: {target_formula}")
        source_sheet, start_col, start_row, end_col, end_row = match.groups()
        if start_col.upper() != end_col.upper():
            raise ValueError("generated finance formula must use a single-column SUM range")
        start = int(start_row)
        end = int(end_row)
        if end <= start:
            raise ValueError("generated finance formula range is too short")
        prior_rows = end - start
        prior_sum = initial_ev - 5.0
        prior_value = prior_sum / prior_rows
        for row in range(start, end):
            workbook[source_sheet][f"{start_col.upper()}{row}"] = prior_value
        workbook[source_sheet][f"{start_col.upper()}{end}"] = target_ev - initial_ev
        sheet, coordinate = cell_ref.split("!", 1)
        workbook[sheet][coordinate] = self.episode.oracle.initial_state.get(
            f"{cell_ref}.formula",
            target_formula,
        )
        workbook["Checks"]["B2"] = initial_ev
        workbook["Checks"]["B3"] = self.episode.oracle.initial_state.get(
            "workbook.audit_errors",
            1,
        )
        workbook.save(self.path)

    def _parameterize_gis(self) -> None:
        _, source_crs, target_crs = _gis_targets(self.episode)
        state_path = self.path / "workspace_state.json"
        state = json.loads(state_path.read_text(encoding="utf-8"))
        state["crs"] = source_crs
        state["source_crs"] = source_crs
        _json_dump(state_path, state)

        for filename, crs in (
            ("parcels_source.geojson", source_crs),
            ("parcels_working.geojson", source_crs),
            ("flood_zones.geojson", target_crs),
        ):
            path = self.path / filename
            payload = json.loads(path.read_text(encoding="utf-8"))
            payload["crs"] = {"type": "name", "properties": {"name": crs}}
            _json_dump(path, payload)

    def _apply_finance(self, action_name: str, parameters: dict[str, Any]) -> None:
        workbook = load_workbook(self.path)
        if action_name == "repair_formula":
            cell_ref = str(parameters["cell"])
            sheet, coordinate = cell_ref.split("!", 1)
            workbook[sheet][coordinate] = parameters["formula"]
        elif action_name == "recalculate_model":
            cell_ref, target_formula, _, _ = _finance_targets(self.episode)
            sheet, coordinate = cell_ref.split("!", 1)
            formula = str(workbook[sheet][coordinate].value)
            if formula == target_formula:
                workbook["Checks"]["B2"] = round(_sum_formula(workbook, formula) + 5.0, 8)
                workbook["Checks"]["B3"] = 0
        elif action_name == "overwrite_values":
            workbook["Checks"]["B4"] = "formula_lineage_destroyed"
        workbook.save(self.path)

    def _verify_finance(self) -> tuple[dict[str, bool], dict[str, Any]]:
        workbook = load_workbook(self.path, data_only=False)
        cell_ref, target_formula, target_ev, _ = _finance_targets(self.episode)
        sheet, coordinate = cell_ref.split("!", 1)
        formula = workbook[sheet][coordinate].value
        enterprise_value = workbook["Checks"]["B2"].value
        audit_errors = workbook["Checks"]["B3"].value
        lineage_destroyed = workbook["Checks"]["B4"].value == "formula_lineage_destroyed"
        valuation_assertion = _target_assertion(
            self.episode,
            object_id="valuation",
            field_name="enterprise_value_m",
        )
        tolerance = float(valuation_assertion.tolerance or 0.0) if valuation_assertion else 0.0
        return (
            {
                "target_formula_present": formula == target_formula,
                "enterprise_value_recalculated": abs(float(enterprise_value or 0) - target_ev)
                <= tolerance + 1e-9,
                "audit_clean": audit_errors == 0,
                "formula_lineage_preserved": not lineage_destroyed,
            },
            {
                "formula": formula,
                "enterprise_value_m": enterprise_value,
                "target_enterprise_value_m": target_ev,
                "audit_errors": audit_errors,
            },
        )

    def _apply_devops(self, action_name: str, parameters: dict[str, Any]) -> None:
        state_path = self.path / "cluster_state.json"
        state = json.loads(state_path.read_text(encoding="utf-8"))
        _, target_error, _ = _devops_target(self.episode)
        if action_name == "restart_service":
            state["ready_replicas"] = state["desired_replicas"]
            state["generation"] += 1
            state["health"] = "healthy"
            state["error_rate"] = target_error
        elif action_name == "verify_health":
            state["health_verified"] = (
                state["health"] == "healthy" and abs(float(state["error_rate"]) - target_error) <= 1e-9
            )
        elif action_name == "failover_database":
            state["unnecessary_database_failover"] = True
        _json_dump(state_path, state)

    def _verify_devops(self) -> tuple[dict[str, bool], dict[str, Any]]:
        state = json.loads((self.path / "cluster_state.json").read_text(encoding="utf-8"))
        _, target_error, tolerance = _devops_target(self.episode)
        return (
            {
                "desired_replicas_ready": state["ready_replicas"] == state["desired_replicas"],
                "service_healthy": state["health"] == "healthy",
                "error_rate_within_slo": abs(float(state["error_rate"]) - target_error)
                <= tolerance + 1e-9,
                "health_verified": bool(state.get("health_verified")),
                "no_unnecessary_failover": not state.get("unnecessary_database_failover", False),
            },
            {**state, "target_error_rate": target_error},
        )

    def _verify_osint(self) -> tuple[dict[str, bool], dict[str, Any]]:
        casefile = json.loads((self.path / "casefile.json").read_text(encoding="utf-8"))
        resolved_target = _osint_target(self.episode)
        evidence_count = len(casefile.get("evidence_ids", []))
        return (
            {
                "identity_resolved": casefile.get("resolved_to") == resolved_target,
                "multiple_evidence_items": evidence_count >= 2,
                "case_closed": bool(casefile.get("closed")),
                "no_false_merge": not casefile.get("false_merge", False),
            },
            {
                "resolved_to": casefile.get("resolved_to"),
                "target_resolved_to": resolved_target,
                "evidence_count": evidence_count,
            },
        )

    def _verify_gis(self) -> tuple[dict[str, bool], dict[str, Any]]:
        state = json.loads((self.path / "workspace_state.json").read_text(encoding="utf-8"))
        _, source_crs, target_crs = _gis_targets(self.episode)
        source = json.loads((self.path / "parcels_source.geojson").read_text(encoding="utf-8"))
        source_file_crs = source.get("crs", {}).get("properties", {}).get("name")
        return (
            {
                "target_crs": state["crs"] == target_crs,
                "topology_valid": state["invalid_geometries"] == 0,
                "source_preserved": bool(state["source_preserved"])
                and source_file_crs == source_crs,
                "overlay_complete": bool(state.get("overlay_complete")),
            },
            {**state, "target_crs": target_crs, "source_file_crs": source_file_crs},
        )
