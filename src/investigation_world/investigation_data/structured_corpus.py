from __future__ import annotations

import csv
import hashlib
import json
from datetime import date, datetime
from enum import Enum
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook  # type: ignore[import-untyped]
from pydantic import Field, model_validator

from .models import (
    AcquisitionArtifact,
    AcquisitionPolicy,
    AIUsePolicy,
    ArtifactClass,
    RedistributionPolicy,
    SourceCatalog,
    SourceSpec,
    StrictModel,
)


class StructuredCorpusError(ValueError):
    """Raised when structured evidence cannot be compiled without semantic loss."""


class StructuredInputFormat(str, Enum):
    CSV = "csv"
    JSON = "json"
    JSONL = "jsonl"
    XLSX = "xlsx"


class FieldExposure(str, Enum):
    PUBLIC = "public"
    VERIFIER = "verifier"
    IGNORE = "ignore"


class CorpusSplit(str, Enum):
    TRAIN_REFERENCE = "train_reference"
    CALIBRATION = "calibration"
    HOLDOUT_CANDIDATE = "holdout_candidate"


ReviewScope = Literal["acquisition", "redistribution", "ai_use", "redaction"]


class StructuredRightsReviewEvidence(StrictModel):
    review_id: str = Field(min_length=1)
    source_id: str = Field(min_length=1)
    source_artifact_id: str = Field(min_length=1)
    scopes: tuple[ReviewScope, ...] = Field(min_length=1)
    policy_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    artifact_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")

    @model_validator(mode="after")
    def validate_review(self) -> "StructuredRightsReviewEvidence":
        if self.review_id != self.review_id.strip():
            raise ValueError("review_id must not contain leading or trailing whitespace")
        if len(self.scopes) != len(set(self.scopes)):
            raise ValueError("rights review scopes must be unique")
        return self


class StructuredFieldRule(StrictModel):
    source_field: str = Field(min_length=1)
    exposure: FieldExposure
    target_field: str | None = None
    required: bool = False

    @model_validator(mode="after")
    def validate_target(self) -> "StructuredFieldRule":
        if self.exposure is FieldExposure.IGNORE:
            if self.target_field is not None:
                raise ValueError("ignored fields may not define target_field")
            return self
        target = self.target_field or self.source_field
        object.__setattr__(self, "target_field", target)
        return self


class StructuredSourceProfile(StrictModel):
    schema_version: Literal["1.0"] = "1.0"
    profile_id: str = Field(min_length=1)
    version: str = Field(min_length=1)
    source_id: str = Field(min_length=1)
    source_artifact_id: str = Field(min_length=1)
    rights_review_id: str | None = None
    input_format: StructuredInputFormat
    domain: str = Field(min_length=1)
    objective: str = Field(min_length=1)
    split: CorpusSplit = CorpusSplit.TRAIN_REFERENCE
    sheet_name: str | None = None
    json_records_key: str | None = None
    source_case_id_fields: tuple[str, ...] = Field(min_length=1)
    title_fields: tuple[str, ...] = Field(min_length=1)
    location_fields: tuple[str, ...] = ()
    event_date_field: str | None = None
    event_date_formats: tuple[str, ...] = ("%Y-%m-%d",)
    event_year_field: str | None = None
    event_month_field: str | None = None
    field_rules: tuple[StructuredFieldRule, ...] = Field(min_length=1)

    @model_validator(mode="after")
    def validate_profile(self) -> "StructuredSourceProfile":
        by_source: dict[str, StructuredFieldRule] = {}
        targets: set[str] = set()
        for rule in self.field_rules:
            if rule.source_field in by_source:
                raise ValueError(f"duplicate source field rule: {rule.source_field}")
            by_source[rule.source_field] = rule
            if rule.target_field is not None:
                if rule.target_field in targets:
                    raise ValueError(f"duplicate target field: {rule.target_field}")
                targets.add(rule.target_field)

        if self.event_date_field is None and self.event_year_field is None:
            raise ValueError("event_date_field or event_year_field is required")

        structural = (
            *self.title_fields,
            *self.location_fields,
            self.event_date_field,
            self.event_year_field,
            self.event_month_field,
        )
        for field_name in (item for item in structural if item is not None):
            structural_rule = by_source.get(field_name)
            if structural_rule is None:
                raise ValueError(f"structural field is not classified: {field_name}")
            if structural_rule.exposure is not FieldExposure.PUBLIC:
                raise ValueError(f"structural field must be public: {field_name}")

        for field_name in self.source_case_id_fields:
            identity_rule = by_source.get(field_name)
            if identity_rule is None:
                raise ValueError(f"source identity field is not classified: {field_name}")
            if identity_rule.exposure is FieldExposure.VERIFIER:
                raise ValueError(
                    f"source identity field cannot depend on verifier-only data: {field_name}"
                )

        if not any(rule.exposure is FieldExposure.VERIFIER for rule in self.field_rules):
            raise ValueError("structured source profile must define at least one verifier field")
        if self.rights_review_id is not None:
            if not self.rights_review_id.strip():
                raise ValueError("rights_review_id must be non-empty when supplied")
            if self.rights_review_id != self.rights_review_id.strip():
                raise ValueError(
                    "rights_review_id must not contain leading or trailing whitespace"
                )
        return self


class StructuredInvestigationCase(StrictModel):
    case_id: str
    source_id: str
    source_artifact_id: str
    source_identity: tuple[str, ...]
    row_number: int = Field(ge=1)
    title: str
    domain: str
    event_date: date
    location: str | None = None
    split: CorpusSplit
    objective: str
    public_payload: dict[str, Any]
    verifier_payload: dict[str, Any]

    def public_projection(self) -> dict[str, Any]:
        return {
            "case_id": self.case_id,
            "source_id": self.source_id,
            "source_artifact_id": self.source_artifact_id,
            "title": self.title,
            "domain": self.domain,
            "event_date": self.event_date.isoformat(),
            "location": self.location,
            "split": self.split.value,
            "objective": self.objective,
            "evidence": self.public_payload,
        }

    def verifier_projection(self) -> dict[str, Any]:
        return {
            "case_id": self.case_id,
            "source_id": self.source_id,
            "source_artifact_id": self.source_artifact_id,
            "source_identity": self.source_identity,
            "row_number": self.row_number,
            "verifier": self.verifier_payload,
        }


class StructuredInvestigationCorpus(StrictModel):
    schema_version: Literal["1.0"] = "1.0"
    dataset_id: str
    version: str
    as_of: date
    profile_id: str
    source_id: str
    source_artifact_id: str
    source_artifact_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    catalog_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    redistribution_policy: RedistributionPolicy
    license_expression: str
    terms_url: str
    attribution_required: bool
    rights_review: StructuredRightsReviewEvidence | None = None
    cases: tuple[StructuredInvestigationCase, ...]

    def public_hash(self) -> str:
        return _stable_hash(tuple(case.public_projection() for case in self.cases))

    def verifier_hash(self) -> str:
        return _stable_hash(tuple(case.verifier_projection() for case in self.cases))


class StructuredCorpusWriteResult(StrictModel):
    dataset_id: str
    cases: int = Field(ge=0)
    public_output: str
    manifest_output: str
    public_hash: str = Field(pattern=r"^[0-9a-f]{64}$")
    verifier_output: str | None = None
    verifier_hash: str | None = Field(default=None, pattern=r"^[0-9a-f]{64}$")


def _stable_hash(value: Any) -> str:
    payload = json.dumps(
        value,
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=False,
        allow_nan=False,
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while True:
            chunk = handle.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def _catalog_sha256(catalog: SourceCatalog) -> str:
    return _stable_hash(catalog.model_dump(mode="json"))


def _find_source(catalog: SourceCatalog, source_id: str) -> SourceSpec:
    for source in catalog.sources:
        if source.source_id == source_id:
            return source
    raise StructuredCorpusError(f"source is absent from canonical catalog: {source_id}")


def _find_artifact(source: SourceSpec, artifact_id: str) -> AcquisitionArtifact:
    artifact = next((item for item in source.artifacts if item.artifact_id == artifact_id), None)
    if artifact is None:
        raise StructuredCorpusError(
            f"artifact {artifact_id!r} is not declared by canonical source {source.source_id!r}"
        )
    return artifact


def _review_policy_payload(source: SourceSpec) -> dict[str, Any]:
    return {
        "rights": source.rights.model_dump(mode="json"),
        "requires_redaction_review": source.requires_redaction_review,
    }


def _review_policy_sha256(source: SourceSpec) -> str:
    return _stable_hash(_review_policy_payload(source))


def _artifact_authority_sha256(artifact: AcquisitionArtifact) -> str:
    return _stable_hash(artifact.model_dump(mode="json"))


def _required_review_scopes(source: SourceSpec) -> tuple[ReviewScope, ...]:
    scopes: list[ReviewScope] = []
    if source.rights.acquisition is AcquisitionPolicy.REVIEW_REQUIRED:
        scopes.append("acquisition")
    if source.rights.redistribution is RedistributionPolicy.REVIEW_REQUIRED:
        scopes.append("redistribution")
    if source.rights.ai_use in {
        AIUsePolicy.REVIEW_REQUIRED,
        AIUsePolicy.ALLOWED_WITH_CONDITIONS,
    }:
        scopes.append("ai_use")
    if source.requires_redaction_review:
        scopes.append("redaction")
    return tuple(scopes)


def _validate_review_evidence(
    profile: StructuredSourceProfile,
    source: SourceSpec,
    artifact: AcquisitionArtifact,
    rights_reviews: tuple[StructuredRightsReviewEvidence, ...],
) -> StructuredRightsReviewEvidence | None:
    required_scopes = _required_review_scopes(source)
    review_id = profile.rights_review_id

    if not required_scopes:
        if review_id is not None:
            raise StructuredCorpusError(
                f"source {source.source_id!r} requires no rights/redaction review; "
                "unsolicited rights_review_id is not authoritative"
            )
        return None

    if review_id is None:
        raise StructuredCorpusError(
            f"source {source.source_id!r} requires scoped rights/redaction review "
            "before public corpus use"
        )

    matches = [evidence for evidence in rights_reviews if evidence.review_id == review_id]
    if len(matches) != 1:
        raise StructuredCorpusError(
            f"rights_review_id {review_id!r} must resolve to exactly one supplied "
            f"review evidence record; found {len(matches)}"
        )
    evidence = matches[0]
    if evidence.source_id != source.source_id:
        raise StructuredCorpusError(
            f"rights review {evidence.review_id!r} is scoped to source "
            f"{evidence.source_id!r}, not {source.source_id!r}"
        )
    if evidence.source_artifact_id != profile.source_artifact_id:
        raise StructuredCorpusError(
            f"rights review {evidence.review_id!r} is scoped to artifact "
            f"{evidence.source_artifact_id!r}, not {profile.source_artifact_id!r}"
        )
    if set(evidence.scopes) != set(required_scopes):
        raise StructuredCorpusError(
            f"rights review {evidence.review_id!r} has scopes {sorted(evidence.scopes)!r}; "
            f"required scopes are {sorted(required_scopes)!r}"
        )
    expected_policy_sha256 = _review_policy_sha256(source)
    if evidence.policy_sha256 != expected_policy_sha256:
        raise StructuredCorpusError(
            f"rights review {evidence.review_id!r} does not bind the current canonical "
            "rights/redaction policy"
        )
    expected_artifact_sha256 = _artifact_authority_sha256(artifact)
    if evidence.artifact_sha256 != expected_artifact_sha256:
        raise StructuredCorpusError(
            f"rights review {evidence.review_id!r} does not bind the current canonical "
            "artifact definition"
        )
    return evidence


def _validate_policy(
    profile: StructuredSourceProfile,
    catalog: SourceCatalog,
    rights_reviews: tuple[StructuredRightsReviewEvidence, ...],
) -> tuple[SourceSpec, AcquisitionArtifact, StructuredRightsReviewEvidence | None]:
    source = _find_source(catalog, profile.source_id)
    artifact = _find_artifact(source, profile.source_artifact_id)

    if source.rights.acquisition is AcquisitionPolicy.BLOCKED:
        raise StructuredCorpusError(f"source {source.source_id!r} is blocked for acquisition")
    if source.rights.ai_use is AIUsePolicy.BLOCKED:
        raise StructuredCorpusError(f"source {source.source_id!r} is blocked for AI use")
    if source.rights.redistribution is RedistributionPolicy.BLOCKED:
        raise StructuredCorpusError(
            f"source {source.source_id!r} is blocked for public redistribution"
        )
    if (
        source.rights.redistribution is RedistributionPolicy.ATTRIBUTION_REQUIRED
        and not source.rights.attribution_required
    ):
        raise StructuredCorpusError(
            f"source {source.source_id!r} has inconsistent attribution policy"
        )
    if (
        source.rights.acquisition is AcquisitionPolicy.METADATA_ONLY
        and artifact.artifact_class is not ArtifactClass.METADATA
    ):
        raise StructuredCorpusError(
            f"source {source.source_id!r} permits metadata-only acquisition, but artifact "
            f"{artifact.artifact_id!r} is {artifact.artifact_class.value!r}"
        )

    evidence = _validate_review_evidence(profile, source, artifact, rights_reviews)
    return source, artifact, evidence


def _normalize_value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, dict):
        return {str(key): _normalize_value(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_normalize_value(item) for item in value]
    return str(value)


def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise StructuredCorpusError("CSV input has no header row")
        if len(reader.fieldnames) != len(set(reader.fieldnames)):
            raise StructuredCorpusError("CSV input contains duplicate header names")
        return [dict(row) for row in reader]


def _read_json(path: Path, records_key: str | None) -> list[dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if records_key is not None:
        if not isinstance(payload, dict):
            raise StructuredCorpusError("JSON records_key requires an object root")
        payload = payload.get(records_key)
    if not isinstance(payload, list) or not all(isinstance(row, dict) for row in payload):
        raise StructuredCorpusError("JSON structured input must contain a list of objects")
    return [dict(row) for row in payload]


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for line_number, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        if not line.strip():
            continue
        payload = json.loads(line)
        if not isinstance(payload, dict):
            raise StructuredCorpusError(f"JSONL line {line_number} is not an object")
        rows.append(dict(payload))
    return rows


def _read_xlsx(path: Path, sheet_name: str | None) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name is not None else workbook.active
        if worksheet is None:
            raise StructuredCorpusError("XLSX input has no active worksheet")
        iterator = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(iterator)
        except StopIteration as exc:
            raise StructuredCorpusError("XLSX input has no header row") from exc
        headers = [str(value).strip() if value is not None else "" for value in header_row]
        if not headers or any(not header for header in headers):
            raise StructuredCorpusError("XLSX input contains blank header cells")
        if len(headers) != len(set(headers)):
            raise StructuredCorpusError("XLSX input contains duplicate header names")
        return [
            {
                header: _normalize_value(value)
                for header, value in zip(headers, values, strict=False)
            }
            for values in iterator
        ]
    finally:
        workbook.close()


def read_structured_records(
    path: Path,
    profile: StructuredSourceProfile,
) -> list[dict[str, Any]]:
    if profile.input_format is StructuredInputFormat.CSV:
        return _read_csv(path)
    if profile.input_format is StructuredInputFormat.JSON:
        return _read_json(path, profile.json_records_key)
    if profile.input_format is StructuredInputFormat.JSONL:
        return _read_jsonl(path)
    if profile.input_format is StructuredInputFormat.XLSX:
        return _read_xlsx(path, profile.sheet_name)
    raise StructuredCorpusError(f"unsupported structured input format: {profile.input_format}")


def _nonempty(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def _parse_event_date(row: dict[str, Any], profile: StructuredSourceProfile) -> date:
    if profile.event_date_field is not None:
        raw = row.get(profile.event_date_field)
        if not _nonempty(raw):
            raise StructuredCorpusError(
                f"missing event date field: {profile.event_date_field}"
            )
        text = str(raw).strip()
        for format_string in profile.event_date_formats:
            try:
                return datetime.strptime(text, format_string).date()
            except ValueError:
                continue
        try:
            return date.fromisoformat(text[:10])
        except ValueError as exc:
            raise StructuredCorpusError(f"could not parse event date {text!r}") from exc

    assert profile.event_year_field is not None
    year_raw = row.get(profile.event_year_field)
    if not _nonempty(year_raw):
        raise StructuredCorpusError(
            f"missing event year field: {profile.event_year_field}"
        )
    year = int(float(str(year_raw)))
    month = 1
    if profile.event_month_field is not None:
        month_raw = row.get(profile.event_month_field)
        if _nonempty(month_raw):
            month = int(float(str(month_raw)))
    if month < 1 or month > 12:
        raise StructuredCorpusError(f"invalid event month: {month}")
    return date(year, month, 1)


def _joined_fields(row: dict[str, Any], fields: tuple[str, ...]) -> str | None:
    values = [str(row.get(field, "")).strip() for field in fields]
    nonempty = [value for value in values if value]
    return " — ".join(nonempty) if nonempty else None


def _source_identity(
    row: dict[str, Any],
    profile: StructuredSourceProfile,
) -> tuple[str, ...]:
    values = tuple(str(row.get(field, "")).strip() for field in profile.source_case_id_fields)
    if any(not value for value in values):
        raise StructuredCorpusError("source case identity fields must all be non-empty")
    return values


def _classify_row(
    row: dict[str, Any],
    profile: StructuredSourceProfile,
    *,
    row_number: int,
) -> StructuredInvestigationCase:
    rule_by_field = {rule.source_field: rule for rule in profile.field_rules}
    unmapped = sorted(field for field in row if field not in rule_by_field)
    if unmapped:
        raise StructuredCorpusError(
            f"row {row_number} contains unclassified fields: {', '.join(unmapped)}"
        )

    public_payload: dict[str, Any] = {}
    verifier_payload: dict[str, Any] = {}
    for rule in profile.field_rules:
        value = _normalize_value(row.get(rule.source_field))
        if rule.required and not _nonempty(value):
            raise StructuredCorpusError(
                f"row {row_number} missing required field: {rule.source_field}"
            )
        if rule.exposure is FieldExposure.IGNORE:
            continue
        assert rule.target_field is not None
        if rule.exposure is FieldExposure.PUBLIC:
            public_payload[rule.target_field] = value
        else:
            verifier_payload[rule.target_field] = value

    identity = _source_identity(row, profile)
    case_id = f"case-{_stable_hash((profile.source_id, identity))[:24]}"
    title = _joined_fields(row, profile.title_fields)
    if title is None:
        raise StructuredCorpusError(f"row {row_number} has no usable title fields")
    return StructuredInvestigationCase(
        case_id=case_id,
        source_id=profile.source_id,
        source_artifact_id=profile.source_artifact_id,
        source_identity=identity,
        row_number=row_number,
        title=title,
        domain=profile.domain,
        event_date=_parse_event_date(row, profile),
        location=_joined_fields(row, profile.location_fields),
        split=profile.split,
        objective=profile.objective,
        public_payload=public_payload,
        verifier_payload=verifier_payload,
    )


def compile_structured_investigation_corpus(
    profile: StructuredSourceProfile,
    input_path: Path,
    catalog: SourceCatalog,
    *,
    dataset_id: str,
    version: str,
    as_of: date,
    rights_reviews: tuple[StructuredRightsReviewEvidence, ...] = (),
) -> StructuredInvestigationCorpus:
    source, artifact, rights_review = _validate_policy(profile, catalog, rights_reviews)

    source_artifact_sha256 = _file_sha256(input_path)
    if (
        artifact.expected_sha256 is not None
        and source_artifact_sha256 != artifact.expected_sha256
    ):
        raise StructuredCorpusError(
            f"local artifact SHA-256 {source_artifact_sha256} does not match canonical "
            f"expected_sha256 {artifact.expected_sha256} for "
            f"{source.source_id}/{artifact.artifact_id}"
        )

    rows = read_structured_records(input_path, profile)
    cases = tuple(
        _classify_row(row, profile, row_number=index)
        for index, row in enumerate(rows, start=1)
    )
    case_ids = [case.case_id for case in cases]
    if len(case_ids) != len(set(case_ids)):
        raise StructuredCorpusError("duplicate canonical source-case identity")
    return StructuredInvestigationCorpus(
        dataset_id=dataset_id,
        version=version,
        as_of=as_of,
        profile_id=profile.profile_id,
        source_id=profile.source_id,
        source_artifact_id=profile.source_artifact_id,
        source_artifact_sha256=source_artifact_sha256,
        catalog_sha256=_catalog_sha256(catalog),
        redistribution_policy=source.rights.redistribution,
        license_expression=source.rights.license_expression,
        terms_url=source.rights.terms_url,
        attribution_required=source.rights.attribution_required,
        rights_review=rights_review,
        cases=cases,
    )


def audit_structured_investigation_corpus(
    corpus: StructuredInvestigationCorpus,
    profile: StructuredSourceProfile,
) -> dict[str, Any]:
    public_targets = {
        rule.target_field
        for rule in profile.field_rules
        if rule.exposure is FieldExposure.PUBLIC and rule.target_field is not None
    }
    verifier_targets = {
        rule.target_field
        for rule in profile.field_rules
        if rule.exposure is FieldExposure.VERIFIER and rule.target_field is not None
    }
    overlap = sorted(public_targets & verifier_targets)
    duplicate_ids = len(corpus.cases) - len({case.case_id for case in corpus.cases})
    leaked_cases = [
        case.case_id
        for case in corpus.cases
        if verifier_targets.intersection(case.public_payload)
    ]
    source_identity_exposed = any(
        "source_identity" in case.public_projection() or "row_number" in case.public_projection()
        for case in corpus.cases
    )
    passed = not overlap and duplicate_ids == 0 and not leaked_cases and not source_identity_exposed
    return {
        "passed": passed,
        "cases": len(corpus.cases),
        "public_fields": sorted(public_targets),
        "verifier_fields": sorted(verifier_targets),
        "field_overlap": overlap,
        "duplicate_case_ids": duplicate_ids,
        "leaked_case_ids": leaked_cases,
        "source_identity_exposed": source_identity_exposed,
        "public_hash": corpus.public_hash(),
    }


def _validate_distinct_output_paths(
    *,
    public_output: Path,
    manifest_output: Path,
    verifier_output: Path | None,
) -> None:
    outputs = [("public_output", public_output), ("manifest_output", manifest_output)]
    if verifier_output is not None:
        outputs.append(("verifier_output", verifier_output))

    resolved: dict[Path, str] = {}
    for label, path in outputs:
        try:
            canonical = path.resolve()
        except OSError as exc:
            raise StructuredCorpusError(f"could not resolve {label}: {path}") from exc
        previous = resolved.get(canonical)
        if previous is not None:
            raise StructuredCorpusError(
                f"output paths must be distinct: {previous} and {label} "
                f"resolve to {canonical}"
            )
        resolved[canonical] = label

    for index, (left_label, left_path) in enumerate(outputs):
        if not left_path.exists():
            continue
        for right_label, right_path in outputs[index + 1 :]:
            if not right_path.exists():
                continue
            try:
                aliases = left_path.samefile(right_path)
            except OSError as exc:
                raise StructuredCorpusError(
                    "could not verify output path separation: "
                    f"{left_label}, {right_label}"
                ) from exc
            if aliases:
                raise StructuredCorpusError(
                    f"output paths must be distinct: {left_label} and {right_label} "
                    "alias the same file"
                )


def write_structured_investigation_corpus(
    corpus: StructuredInvestigationCorpus,
    profile: StructuredSourceProfile,
    *,
    public_output: Path,
    manifest_output: Path,
    verifier_output: Path | None = None,
) -> StructuredCorpusWriteResult:
    audit = audit_structured_investigation_corpus(corpus, profile)
    if not audit["passed"]:
        raise StructuredCorpusError(f"structured corpus boundary audit failed: {audit}")
    _validate_distinct_output_paths(
        public_output=public_output,
        manifest_output=manifest_output,
        verifier_output=verifier_output,
    )

    public_output.parent.mkdir(parents=True, exist_ok=True)
    manifest_output.parent.mkdir(parents=True, exist_ok=True)
    public_output.write_text(
        "".join(
            json.dumps(case.public_projection(), sort_keys=True) + "\n"
            for case in corpus.cases
        ),
        encoding="utf-8",
    )
    verifier_hash: str | None = None
    verifier_path: str | None = None
    if verifier_output is not None:
        verifier_output.parent.mkdir(parents=True, exist_ok=True)
        verifier_output.write_text(
            "".join(
                json.dumps(case.verifier_projection(), sort_keys=True) + "\n"
                for case in corpus.cases
            ),
            encoding="utf-8",
        )
        verifier_hash = corpus.verifier_hash()
        verifier_path = str(verifier_output)

    review_payload = (
        corpus.rights_review.model_dump(mode="json")
        if corpus.rights_review is not None
        else None
    )
    manifest = {
        "schema_version": corpus.schema_version,
        "dataset_id": corpus.dataset_id,
        "version": corpus.version,
        "as_of": corpus.as_of.isoformat(),
        "profile_id": corpus.profile_id,
        "source_id": corpus.source_id,
        "source_artifact_id": corpus.source_artifact_id,
        "source_artifact_sha256": corpus.source_artifact_sha256,
        "catalog_sha256": corpus.catalog_sha256,
        "rights": {
            "redistribution": corpus.redistribution_policy.value,
            "license_expression": corpus.license_expression,
            "terms_url": corpus.terms_url,
            "attribution_required": corpus.attribution_required,
            "review_id": (
                corpus.rights_review.review_id
                if corpus.rights_review is not None
                else None
            ),
            "review": review_payload,
        },
        "cases": len(corpus.cases),
        "split": profile.split.value,
        "public_hash": corpus.public_hash(),
        "verifier_materialized": verifier_output is not None,
        "audit": {
            "passed": audit["passed"],
            "duplicate_case_ids": audit["duplicate_case_ids"],
            "field_overlap": audit["field_overlap"],
            "leaked_case_count": len(audit["leaked_case_ids"]),
            "source_identity_exposed": audit["source_identity_exposed"],
        },
    }
    manifest_output.write_text(
        json.dumps(manifest, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    return StructuredCorpusWriteResult(
        dataset_id=corpus.dataset_id,
        cases=len(corpus.cases),
        public_output=str(public_output),
        manifest_output=str(manifest_output),
        public_hash=corpus.public_hash(),
        verifier_output=verifier_path,
        verifier_hash=verifier_hash,
    )
