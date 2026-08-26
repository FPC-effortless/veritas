from __future__ import annotations

import random

from openpyxl import load_workbook

from investigation_world.operational import (
    EpisodeSubmission,
    NativeOperationalRuntime,
    OperationalDistributionConfig,
    WorldDomain,
    build_devops_incident_world,
    build_enterprise_operations_world,
    build_financial_spreadsheet_world,
    build_gis_operations_world,
    build_investigation_osint_world,
    compile_operational_distribution,
    validate_operational_distribution,
)
from investigation_world.operational.realism import apply_domain_realism


def _deep_episode(builder, domain: WorldDomain):
    episode = builder(seed=42)
    return apply_domain_realism(
        episode,
        rng=random.Random(9001),
        index=7,
        scenario_family={
            WorldDomain.FINANCIAL_SPREADSHEET: "dcf_formula_repair",
            WorldDomain.ENTERPRISE_OPERATIONS: "discount_control",
            WorldDomain.DEVOPS_INCIDENT_RESPONSE: "service_availability",
            WorldDomain.INVESTIGATION_OSINT: "identity_resolution",
            WorldDomain.GIS_OPERATIONS: "projection_alignment",
        }[domain],
    )


def _params(episode, action_name: str, ordinal: int = 0):
    effects = [effect for effect in episode.oracle.action_effects if effect.action_name == action_name]
    return dict(effects[ordinal].required_parameters)


def _submission(episode):
    return EpisodeSubmission(
        conclusion="completed with native artifact verification",
        evidence_ids=list(episode.oracle.required_evidence_ids),
        confidence=0.95,
    )


def _execute_required_actions(runtime: NativeOperationalRuntime) -> None:
    episode = runtime.episode
    for action in episode.oracle.required_action_order:
        effects = [effect for effect in episode.oracle.action_effects if effect.action_name == action]
        required_count = episode.oracle.required_action_counts.get(action, 1)
        for ordinal in range(required_count):
            effect = effects[min(ordinal, len(effects) - 1)]
            runtime.act(action, **dict(effect.required_parameters))


def test_financial_native_xlsx_execution_and_verification(tmp_path):
    episode = _deep_episode(build_financial_spreadsheet_world, WorldDomain.FINANCIAL_SPREADSHEET)
    runtime = NativeOperationalRuntime(episode, artifact_root=tmp_path)
    path = runtime.materialize_artifact()
    assert path.suffix == ".xlsx"
    assert load_workbook(path)["DCF"]["F18"].value == "=SUM(Revenue!B2:B12)"

    _execute_required_actions(runtime)

    artifact = runtime.artifact_verification()
    assert artifact.valid is True
    result = runtime.submit(_submission(episode))
    assert result.state == 1.0
    assert result.outcome == 1.0
    assert not [item for item in result.process_violations if item.startswith("native_artifact:")]


def test_enterprise_native_sqlite_replica_tracks_cross_system_actions(tmp_path):
    episode = _deep_episode(build_enterprise_operations_world, WorldDomain.ENTERPRISE_OPERATIONS)
    runtime = NativeOperationalRuntime(episode, artifact_root=tmp_path)
    _execute_required_actions(runtime)
    artifact = runtime.artifact_verification()
    assert artifact.valid is True
    assert artifact.measurements["audit_events"] >= 2
    result = runtime.submit(_submission(episode))
    assert result.overall_reward > 0.95


def test_devops_native_declarative_sandbox_reaches_verified_recovery(tmp_path):
    episode = _deep_episode(build_devops_incident_world, WorldDomain.DEVOPS_INCIDENT_RESPONSE)
    runtime = NativeOperationalRuntime(episode, artifact_root=tmp_path)
    _execute_required_actions(runtime)
    artifact = runtime.artifact_verification()
    assert artifact.valid is True
    assert artifact.measurements["ready_replicas"] == artifact.measurements["desired_replicas"]
    assert artifact.measurements["health_verified"] is True


def test_osint_native_corpus_requires_multiple_linked_evidence_items(tmp_path):
    episode = _deep_episode(build_investigation_osint_world, WorldDomain.INVESTIGATION_OSINT)
    runtime = NativeOperationalRuntime(episode, artifact_root=tmp_path)
    _execute_required_actions(runtime)
    artifact = runtime.artifact_verification()
    assert artifact.valid is True
    assert artifact.measurements["evidence_count"] == 2


def test_gis_native_geojson_reprojects_repairs_and_overlays(tmp_path):
    episode = _deep_episode(build_gis_operations_world, WorldDomain.GIS_OPERATIONS)
    runtime = NativeOperationalRuntime(episode, artifact_root=tmp_path)
    _execute_required_actions(runtime)
    artifact = runtime.artifact_verification()
    assert artifact.valid is True
    assert artifact.measurements["crs"] == "EPSG:32631"
    assert artifact.measurements["invalid_geometries"] == 0
    assert artifact.measurements["source_preserved"] is True


def test_native_artifact_failure_is_scored_inside_existing_state_contract(tmp_path):
    episode = _deep_episode(build_financial_spreadsheet_world, WorldDomain.FINANCIAL_SPREADSHEET)
    runtime = NativeOperationalRuntime(episode, artifact_root=tmp_path)
    _execute_required_actions(runtime)

    workbook = load_workbook(runtime.materialize_artifact())
    workbook["DCF"]["F18"] = "=0"
    workbook.save(runtime.materialize_artifact())

    result = runtime.submit(_submission(episode))
    assert result.state < 1.0
    assert result.outcome < 1.0
    assert "native_artifact:target_formula_present" in result.process_violations


def test_production_distribution_keeps_scale_contract_with_lazy_native_descriptors():
    config = OperationalDistributionConfig(
        seed=19,
        train_per_domain=1,
        iid_per_domain=1,
        ood_per_domain=1,
        adversarial_per_domain=1,
    )
    cases = compile_operational_distribution(config)
    validation = validate_operational_distribution(cases, config=config)
    assert validation["valid"] is True, validation["errors"]
    assert len(cases) == 20
    assert validation["manifest"]["version"] == "operational-production-v3"
    for case in cases:
        descriptor = case.episode.task.metadata["native_artifact"]
        assert descriptor["metadata"]["lazy"] is True
        assert descriptor["source_record_ids"] == [record.record_id for record in case.episode.records]


def test_parameterized_train_cases_materialize_and_verify_native_artifacts(tmp_path):
    config = OperationalDistributionConfig(
        seed=73,
        train_per_domain=4,
        iid_per_domain=1,
        ood_per_domain=1,
        adversarial_per_domain=1,
    )
    cases = compile_operational_distribution(config)
    selected = {}
    for case in cases:
        if case.split.value != "train":
            continue
        selected.setdefault(case.episode.task.domain, []).append(case)

    assert set(selected) == set(WorldDomain)
    for domain, domain_cases in selected.items():
        # Use the fourth training case so formula periods, service names,
        # identities, and CRS pairs differ from the reference examples.
        case = domain_cases[3]
        runtime = NativeOperationalRuntime(
            case.episode,
            artifact_root=tmp_path / domain.value,
        )
        _execute_required_actions(runtime)
        artifact = runtime.artifact_verification()
        assert artifact.valid is True, (domain.value, artifact.model_dump())
        result = runtime.submit(_submission(case.episode))
        assert result.state == 1.0, (domain.value, result.model_dump())
        assert result.outcome == 1.0, (domain.value, result.model_dump())
