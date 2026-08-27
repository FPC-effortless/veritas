from __future__ import annotations

import argparse
import json
import shutil
import tempfile
import time
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from collections import Counter
from pathlib import Path
from typing import Any

from investigation_world.operational_world import (
    CalibrationProfile,
    CompanySizeBand,
    IndustryFamily,
    RegionGroup,
    build_bootstrap_calibration,
    compose_calibration_profiles,
)
from investigation_world.operational_world.external_profiles import (
    build_gleif_entity_profile,
    build_retail_transaction_profile,
    build_sec_financial_profile,
)


UCI_RETAIL_URL = "https://archive.ics.uci.edu/static/public/502/online%2Bretail%2Bii.zip"
GLEIF_API = "https://api.gleif.org/api/v1/lei-records"
SEC_FRAME_BASE = "https://data.sec.gov/api/xbrl/frames/us-gaap"
USER_AGENT = "Veritas-Operational-World-Calibration/1.0 (+https://github.com/FPC-effortless/veritas)"

# Equal-target country-stratified GLEIF sampling deliberately spans distinct economic systems.
# It calibrates entity shape only; country frequency in this sample is never interpreted as
# prevalence in the world economy.
GLEIF_SAMPLE_COUNTRIES: tuple[str, ...] = (
    "NG", "GH", "KE", "ZA", "EG", "MA",  # Africa
    "US", "CA", "MX", "BR", "AR", "CL",  # Americas
    "GB", "DE", "FR", "IT", "PL", "NL",  # Europe
    "AE", "SA", "IN", "PK",  # Middle East / South Asia
    "CN", "JP", "SG", "AU",  # East Asia / Pacific
)


def _request_json(url: str, *, timeout: int = 180) -> dict[str, Any]:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": USER_AGENT,
            "Accept": "application/json, application/vnd.api+json",
        },
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return json.load(response)


def _download(url: str, destination: Path, *, timeout: int = 300) -> int:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(request, timeout=timeout) as response, destination.open("wb") as target:
        shutil.copyfileobj(response, target)
    return destination.stat().st_size


def _gleif_country_url(country_code: str, page_size: int) -> str:
    query = urllib.parse.urlencode(
        {
            "filter[entity.legalAddress.country]": country_code,
            "page[number]": 1,
            "page[size]": page_size,
        }
    )
    return f"{GLEIF_API}?{query}"


def materialize_gleif(
    *,
    countries: tuple[str, ...] = GLEIF_SAMPLE_COUNTRIES,
    page_size: int = 200,
) -> tuple[CalibrationProfile, dict[str, Any]]:
    records: list[dict[str, Any]] = []
    country_rows: dict[str, int] = {}
    for index, country_code in enumerate(countries):
        payload = _request_json(_gleif_country_url(country_code, page_size))
        country_records = payload.get("data") or []
        records.extend(country_records)
        country_rows[country_code] = len(country_records)
        if index + 1 < len(countries):
            # Remain below GLEIF's documented public rate limit.
            time.sleep(1.05)

    profile = build_gleif_entity_profile(records)
    profile.notes.append(
        "Materialization uses equal-target country-stratified sampling across multiple regions; country frequency in this sample is not a prevalence estimate."
    )
    return profile, {
        "source_id": "gleif_lei",
        "status": "materialized",
        "url": GLEIF_API,
        "records": len(records),
        "sampled_countries": list(countries),
        "country_rows": country_rows,
        "page_size_per_country": page_size,
        "sampling": "country_stratified_first_page_equal_target_mass",
    }


def _normalize_header(value: Any) -> str:
    return "".join(character for character in str(value or "").casefold() if character.isalnum())


def _header_index(row: tuple[Any, ...]) -> dict[str, int]:
    return {_normalize_header(value): index for index, value in enumerate(row)}


def _pick(indexes: dict[str, int], *names: str) -> int | None:
    for name in names:
        normalized = _normalize_header(name)
        if normalized in indexes:
            return indexes[normalized]
    return None


def materialize_uci_retail(temp_root: Path) -> tuple[CalibrationProfile, dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - workflow installs the optional dependency.
        raise RuntimeError("materializing Online Retail II requires openpyxl") from exc

    archive = temp_root / "online-retail-ii.zip"
    compressed_bytes = _download(UCI_RETAIL_URL, archive)
    with zipfile.ZipFile(archive) as bundle:
        workbook_names = [name for name in bundle.namelist() if name.lower().endswith(".xlsx")]
        if not workbook_names:
            raise RuntimeError("Online Retail II archive did not contain an xlsx workbook")
        workbook_path = temp_root / Path(workbook_names[0]).name
        with bundle.open(workbook_names[0]) as source, workbook_path.open("wb") as target:
            shutil.copyfileobj(source, target)

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    invoice_state: dict[str, dict[str, Any]] = {}
    source_rows = 0
    for worksheet in workbook.worksheets:
        iterator = worksheet.iter_rows(values_only=True)
        header = next(iterator, None)
        if header is None:
            continue
        indexes = _header_index(tuple(header))
        invoice_col = _pick(indexes, "Invoice", "InvoiceNo")
        stock_col = _pick(indexes, "StockCode")
        quantity_col = _pick(indexes, "Quantity")
        price_col = _pick(indexes, "Price", "UnitPrice")
        customer_col = _pick(indexes, "Customer ID", "CustomerID")
        country_col = _pick(indexes, "Country")
        required = {
            "invoice": invoice_col,
            "stock": stock_col,
            "quantity": quantity_col,
            "price": price_col,
        }
        if any(value is None for value in required.values()):
            raise RuntimeError(
                f"unrecognized Online Retail II columns in {worksheet.title}: {list(indexes)}"
            )

        for row in iterator:
            source_rows += 1
            invoice_raw = row[invoice_col] if invoice_col is not None and invoice_col < len(row) else None
            if invoice_raw is None or invoice_raw == "":
                continue
            invoice_id = str(invoice_raw).strip()
            if not invoice_id:
                continue
            state = invoice_state.setdefault(
                invoice_id,
                {
                    "line_count": 0,
                    "products": set(),
                    "value_gbp": 0.0,
                    "cancelled": invoice_id.casefold().startswith("c"),
                    "customer_id": None,
                    "country": None,
                },
            )
            state["line_count"] += 1
            if stock_col is not None and stock_col < len(row):
                stock_value = row[stock_col]
                if stock_value is not None and stock_value != "":
                    state["products"].add(str(stock_value))
            try:
                quantity = float(row[quantity_col]) if quantity_col is not None else 0.0
                price = float(row[price_col]) if price_col is not None else 0.0
            except (TypeError, ValueError):
                quantity = 0.0
                price = 0.0
            state["value_gbp"] += quantity * price
            if customer_col is not None and customer_col < len(row):
                customer_value = row[customer_col]
                if customer_value is not None and customer_value != "":
                    state["customer_id"] = str(customer_value)
            if country_col is not None and country_col < len(row):
                country_value = row[country_col]
                if country_value is not None and country_value != "":
                    state["country"] = str(country_value).strip()
    workbook.close()

    invoices: list[dict[str, Any]] = []
    customer_invoice_counts: Counter[str] = Counter()
    country_counts: Counter[str] = Counter()
    for state in invoice_state.values():
        invoices.append(
            {
                "line_count": state["line_count"],
                "unique_products": len(state["products"]),
                "value_gbp": round(float(state["value_gbp"]), 4),
                "cancelled": bool(state["cancelled"]),
            }
        )
        if state["customer_id"]:
            customer_invoice_counts[str(state["customer_id"])] += 1
        if state["country"]:
            country_counts[str(state["country"])] += 1

    profile = build_retail_transaction_profile(
        invoices,
        customer_invoice_counts=customer_invoice_counts.values(),
        country_counts=country_counts,
    )
    return profile, {
        "source_id": "online_retail_ii",
        "status": "materialized",
        "url": UCI_RETAIL_URL,
        "compressed_bytes": compressed_bytes,
        "source_rows": source_rows,
        "invoices": len(invoices),
        "customers_with_id": len(customer_invoice_counts),
        "countries": len(country_counts),
    }


def _sec_frame(tag: str, period: str) -> dict[str, Any]:
    quoted_tag = urllib.parse.quote(tag, safe="")
    url = f"{SEC_FRAME_BASE}/{quoted_tag}/USD/{period}.json"
    return _request_json(url)


def _first_sec_frame(candidates: tuple[tuple[str, str], ...]) -> tuple[str, str, dict[str, Any]]:
    errors: list[str] = []
    for tag, period in candidates:
        try:
            payload = _sec_frame(tag, period)
        except urllib.error.HTTPError as exc:
            if exc.code == 404:
                errors.append(f"{tag}/{period}:404")
                continue
            raise
        if payload.get("data"):
            return tag, period, payload
        errors.append(f"{tag}/{period}:empty")
    raise RuntimeError("SEC did not return a usable frame: " + ", ".join(errors))


def materialize_sec_finance() -> tuple[CalibrationProfile, dict[str, Any]]:
    asset_tag, instant_period, assets = _first_sec_frame(
        (("Assets", "CY2025Q4I"), ("Assets", "CY2025Q3I"))
    )
    liability_tag, liability_period, liabilities = _first_sec_frame(
        (("Liabilities", instant_period), ("LiabilitiesAndStockholdersEquity", instant_period))
    )
    payable_tag, payable_period, accounts_payable = _first_sec_frame(
        (
            ("AccountsPayableCurrent", instant_period),
            ("AccountsPayableAndAccruedLiabilitiesCurrent", instant_period),
        )
    )
    revenue_tag, revenue_period, revenues = _first_sec_frame(
        (
            ("Revenues", "CY2025"),
            ("RevenueFromContractWithCustomerExcludingAssessedTax", "CY2025"),
            ("SalesRevenueNet", "CY2025"),
        )
    )
    profile = build_sec_financial_profile(
        assets=assets,
        liabilities=liabilities,
        accounts_payable=accounts_payable,
        revenues=revenues,
    )
    return profile, {
        "source_id": "sec_edgar_xbrl",
        "status": "materialized",
        "frames": {
            "assets": {"tag": asset_tag, "period": instant_period, "rows": len(assets.get("data") or [])},
            "liabilities": {"tag": liability_tag, "period": liability_period, "rows": len(liabilities.get("data") or [])},
            "accounts_payable": {"tag": payable_tag, "period": payable_period, "rows": len(accounts_payable.get("data") or [])},
            "revenues": {"tag": revenue_tag, "period": revenue_period, "rows": len(revenues.get("data") or [])},
        },
    }


def _write_profile(profile: CalibrationProfile, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(profile.model_dump_json(indent=2), encoding="utf-8")


def _remove_if_exists(path: Path) -> None:
    if path.exists():
        path.unlink()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Materialize broad public business calibration profiles for Veritas."
    )
    parser.add_argument("--output-dir", type=Path, default=Path("calibration/public/business"))
    parser.add_argument(
        "--gleif-page-size",
        type=int,
        default=200,
        help="Records requested per country; maximum supported by GLEIF is 200.",
    )
    parser.add_argument(
        "--require-sec",
        action="store_true",
        help="Fail if SEC frames cannot be acquired. Default is to record SEC as unavailable and continue.",
    )
    args = parser.parse_args()
    if not 1 <= args.gleif_page_size <= 200:
        raise SystemExit("--gleif-page-size must be between 1 and 200")

    args.output_dir.mkdir(parents=True, exist_ok=True)
    acquisitions: list[dict[str, Any]] = []
    materialized_profiles: list[CalibrationProfile] = []

    with tempfile.TemporaryDirectory(prefix="veritas-business-calibration-") as temp_dir:
        temp_root = Path(temp_dir)
        gleif_profile, gleif_meta = materialize_gleif(page_size=args.gleif_page_size)
        retail_profile, retail_meta = materialize_uci_retail(temp_root)
        acquisitions.extend([gleif_meta, retail_meta])

        sec_profile: CalibrationProfile | None = None
        try:
            sec_profile, sec_meta = materialize_sec_finance()
            acquisitions.append(sec_meta)
        except urllib.error.HTTPError as exc:
            if args.require_sec or exc.code not in {403, 429}:
                raise
            acquisitions.append(
                {
                    "source_id": "sec_edgar_xbrl",
                    "status": "unavailable_from_runner",
                    "http_status": exc.code,
                    "url": SEC_FRAME_BASE,
                    "reason": (
                        "SEC automated-access policy refused the GitHub-hosted runner. "
                        "No SEC-derived profile was emitted or substituted with unsourced data."
                    ),
                }
            )

    gleif_path = args.output_dir / "global-entity-gleif-v1.json"
    retail_path = args.output_dir / "europe-retail-transactions-medium-v1.json"
    sec_path = args.output_dir / "north-america-public-company-finance-large-v1.json"
    _write_profile(gleif_profile, gleif_path)
    _write_profile(retail_profile, retail_path)
    materialized_profiles.extend([gleif_profile, retail_profile])

    # Clean up the obsolete size-specific GLEIF filename if produced by an earlier run.
    _remove_if_exists(args.output_dir / "global-entity-gleif-medium-v1.json")

    europe_retail_bootstrap = build_bootstrap_calibration(
        region=RegionGroup.EUROPE,
        industry=IndustryFamily.RETAIL,
        size_band=CompanySizeBand.MEDIUM,
    )
    europe_retail_composite = compose_calibration_profiles(
        [europe_retail_bootstrap, gleif_profile, retail_profile],
        profile_id="europe-retail-medium-hybrid-v1",
        region=RegionGroup.EUROPE,
        industry=IndustryFamily.RETAIL,
        size_band=CompanySizeBand.MEDIUM,
    )
    _write_profile(europe_retail_composite, args.output_dir / "europe-retail-medium-hybrid-v1.json")
    materialized_profiles.append(europe_retail_composite)

    if sec_profile is not None:
        _write_profile(sec_profile, sec_path)
        materialized_profiles.append(sec_profile)
        north_america_large_bootstrap = build_bootstrap_calibration(
            region=RegionGroup.NORTH_AMERICA,
            industry=IndustryFamily.GENERIC,
            size_band=CompanySizeBand.LARGE,
        )
        north_america_finance_composite = compose_calibration_profiles(
            [north_america_large_bootstrap, gleif_profile, sec_profile],
            profile_id="north-america-large-finance-hybrid-v1",
            region=RegionGroup.NORTH_AMERICA,
            industry=IndustryFamily.GENERIC,
            size_band=CompanySizeBand.LARGE,
        )
        _write_profile(
            north_america_finance_composite,
            args.output_dir / "north-america-large-finance-hybrid-v1.json",
        )
        materialized_profiles.append(north_america_finance_composite)
    else:
        # Never leave stale SEC outputs behind after an acquisition failure.
        _remove_if_exists(sec_path)
        _remove_if_exists(args.output_dir / "north-america-large-finance-hybrid-v1.json")

    manifest = {
        "format": "veritas-business-calibration-materialization-v3",
        "acquisition": acquisitions,
        "profiles": [
            {
                "profile_id": profile.profile_id,
                "state": profile.state,
                "region": profile.region,
                "industry": profile.industry,
                "size_band": profile.size_band,
                "size_scope": profile.size_scope,
                "source_ids": profile.source_ids,
                "empirical_observation_count": profile.empirical_observation_count,
                "metrics": sorted(profile.distributions),
                "categories": sorted(profile.categories),
            }
            for profile in materialized_profiles
        ],
        "quality_policy": [
            "GLEIF contributes size-agnostic legal-entity shape only; country-stratified sampling is not treated as prevalence.",
            "Online Retail II contributes medium European retail transaction shape only.",
            "SEC contributes large/public-company finance only when its official endpoint admits the materialization runner.",
            "A refused source is recorded as unavailable and is never silently replaced with unverified data.",
            "Composite profiles preserve uncovered assumptions as bootstrap priors and therefore remain hybrid.",
        ],
    }
    (args.output_dir / "manifest.json").write_text(
        json.dumps(manifest, indent=2, default=str), encoding="utf-8"
    )
    print(json.dumps(manifest, indent=2, default=str))


if __name__ == "__main__":
    main()
